"""run_chain_gate_eval.py — Extraction-only eval: causal chain → Hard Gate.

Pipeline (no code generation, no OpenROAD execution):
  1. Load action_db.xlsx
  2. For each step prompt:
       a. Bootstrap causal chain extraction (GPT-4.1-mini)
       b. Hard Gate validation against RAGAPIs_structured.csv + OpenROAD source
       c. If hallucinations found → re-extract with corrective feedback (≤2 retries)
  3. Write results to Excel

Excel columns per step:
  prompt | raw_chain | gate_summary | valid_types | rag_miss_types |
  hallucinated_types | rewrites_needed | final_chain | rewrite_feedback

Usage:
  conda activate prompt
  python causal_verifier_4_1/run_chain_gate_eval.py \\
      --openaiKey sk-... \\
      --actionDBPath src_1_reflector/action_db.xlsx \\
      --RAGApiPath RAGData/RAGAPIs_structured.csv \\
      --resultPath result/ \\
      --level 2 --numCases 10
"""

import argparse
import json
import os
import re
import sys
import time
import urllib.request
import urllib.error
from typing import List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

# ── path setup ────────────────────────────────────────────────────────────────
_ROOT       = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_CAUSAL_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _CAUSAL_DIR)

from structured_rag_gate import StructuredRAGGate, GateReport, TypeStatus

# Reuse chain extraction helpers from run_causal_agent
from run_causal_agent import (
    CHAIN_SYSTEM_PROMPT,
    _call_openai,
    _parse_multi_path_chain,
)


# ─────────────────────────────────────────────────────────────────────────────
# Case loader  (same as run_causal_agent_sequential.py)
# ─────────────────────────────────────────────────────────────────────────────

def load_cases(path: str, level: int,
               sheet: Optional[str] = None) -> List[Tuple[str, List[str]]]:
    import openpyxl
    wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws   = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    cases = []
    for row in rows[1:]:
        if not row[0] or not str(row[0]).strip():
            continue
        complex_prompt = str(row[0]).strip()
        steps = []
        for col_idx in range(1, level + 1):
            val = row[col_idx] if len(row) > col_idx else None
            steps.append(str(val).strip() if val and str(val).strip() else "")
        cases.append((complex_prompt, steps))
    return cases


# ─────────────────────────────────────────────────────────────────────────────
# Extraction with optional corrective feedback (max retries)
# ─────────────────────────────────────────────────────────────────────────────

def extract_chain(task: str, api_key: str, model: str,
                  feedback: str = "") -> Tuple[List[List[str]], str]:
    """Run chain extraction. If feedback is non-empty, prepend it as a correction.

    Returns (paths, action_node).
    """
    user_content = f"Task: {task}"
    if feedback:
        user_content = (
            f"CORRECTION REQUIRED:\n{feedback}\n\n"
            f"Now re-extract the causal chain, fixing the invalid types.\n\n"
            f"Task: {task}"
        )

    text = _call_openai(
        messages=[
            {"role": "system", "content": CHAIN_SYSTEM_PROMPT},
            {"role": "user",   "content": user_content},
        ],
        api_key=api_key,
        model=model,
        max_tokens=300,
    )
    if not text:
        return [], ""
    return _parse_multi_path_chain(text)


def paths_to_str(paths: List[List[str]]) -> str:
    """Human-readable chain string, e.g. Path1: A→B→C | Path2: A→D"""
    parts = []
    for i, path in enumerate(paths, 1):
        parts.append(f"P{i}: " + " → ".join(path))
    return " | ".join(parts)


# ─────────────────────────────────────────────────────────────────────────────
# Main eval loop
# ─────────────────────────────────────────────────────────────────────────────

MAX_REWRITES = 2   # max re-extraction attempts on hallucination

def _clean(s: object) -> str:
    return ILLEGAL_CHARACTERS_RE.sub("", str(s or ""))


def Run(
    action_db_path:  str,
    result_path:     str,
    rag_api_path:    str,   # RAGAPIs_structured.csv
    api_key:         str,
    openai_model:    str           = "gpt-4.1-mini",
    level:           int           = 2,
    num_cases:       Optional[int] = None,
    start_case:      Optional[int] = None,
    cases:           Optional[List[int]] = None,
    run_tag:         str           = "",
    sheet:           Optional[str] = None,
):
    # ── Gate init ─────────────────────────────────────────────────────────────
    print(f"\n[Init] Loading StructuredRAGGate from {os.path.basename(rag_api_path)}...",
          flush=True)
    gate = StructuredRAGGate(rag_api_path)

    # ── Load cases ────────────────────────────────────────────────────────────
    all_cases = load_cases(action_db_path, level, sheet=sheet)
    if cases is not None:
        indexed_cases = [(i, all_cases[i][0], all_cases[i][1])
                         for i in cases if i < len(all_cases)]
    else:
        if start_case is not None:
            all_cases = all_cases[start_case:]
        if num_cases is not None:
            all_cases = all_cases[:num_cases]
        offset = start_case or 0
        indexed_cases = [(offset + i, cp, steps)
                         for i, (cp, steps) in enumerate(all_cases)]

    print(f"[Init] Level={level}, running {len(indexed_cases)} case(s).", flush=True)

    # ── Result workbook ───────────────────────────────────────────────────────
    os.makedirs(result_path, exist_ok=True)
    dataset_tag  = os.path.splitext(os.path.basename(action_db_path))[0]
    tag_suffix   = f"_{run_tag}" if run_tag else ""
    result_file  = os.path.join(
        result_path,
        f"{openai_model.replace('.', '-')}__gate_L{level}_{dataset_tag}{tag_suffix}.xlsx"
    )

    if os.path.exists(result_file):
        wb = load_workbook(result_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Gate Results"

        # Build headers
        headers = {1: "complex_prompt"}
        col = 2
        for n in range(1, level + 1):
            for field in (
                "prompt", "raw_chain",
                "gate_summary",
                "valid_types", "rag_miss_types", "hallucinated_types",
                "rewrites_done", "final_chain",
                "rewrite_feedback",
            ):
                headers[col] = f"step{n}_{field}"
                col += 1
        headers[col] = "overall_gate"
        for c, h in headers.items():
            ws.cell(row=1, column=c, value=h)

    _FIELDS = [
        "prompt", "raw_chain",
        "gate_summary",
        "valid_types", "rag_miss_types", "hallucinated_types",
        "rewrites_done", "final_chain",
        "rewrite_feedback",
    ]

    def _col(step_n: int, field: str) -> int:
        return (step_n - 1) * len(_FIELDS) + _FIELDS.index(field) + 2

    def _overall_col() -> int:
        return level * len(_FIELDS) + 2

    # ── Case loop ─────────────────────────────────────────────────────────────
    total_steps  = 0
    clean_steps  = 0   # steps with no hallucination after rewrite

    for case_i, complex_prompt, steps in indexed_cases:
        row = case_i + 2
        print(f"\n{'='*60}", flush=True)
        print(f"CASE {case_i+1}  (level={level}, {len(steps)} steps)", flush=True)

        ws.cell(row=row, column=1, value=_clean(complex_prompt))

        step_overall: List[str] = []

        for step_idx, raw_prompt in enumerate(steps):
            step_n = step_idx + 1
            if not raw_prompt:
                continue

            total_steps += 1
            print(f"\n  ── Step {step_n}: {raw_prompt[:120]}", flush=True)

            # ── Extract chain ──────────────────────────────────────────────
            paths, action_node = extract_chain(
                raw_prompt, api_key=api_key, model=openai_model
            )
            raw_chain_str = paths_to_str(paths)
            print(f"  [extract] {raw_chain_str or '[FAILED]'}", flush=True)

            # ── Gate validation + optional re-extraction ───────────────────
            gate_report   = gate.validate(paths, task=raw_prompt)
            rewrites_done = 0

            while gate_report.had_hallucinations and rewrites_done < MAX_REWRITES:
                rewrites_done += 1
                print(f"  [gate] HALLUCINATION detected — rewrite #{rewrites_done}", flush=True)
                print(f"  [gate] feedback:\n{gate_report.rewrite_feedback}", flush=True)

                new_paths, new_action = extract_chain(
                    raw_prompt,
                    api_key=api_key, model=openai_model,
                    feedback=gate_report.rewrite_feedback,
                )
                if new_paths:
                    paths       = new_paths
                    action_node = new_action
                    gate_report = gate.validate(paths, task=raw_prompt)
                else:
                    print("  [gate] re-extraction returned empty — keeping original",
                          flush=True)
                    break

            final_chain_str = paths_to_str(paths)
            summary         = gate_report.summary()

            # Collect type lists for output
            valid_types = []
            rag_misses  = []
            hallucinated = []
            for path_reports in gate_report.path_reports:
                for tr in path_reports:
                    if tr.status == TypeStatus.VALID:
                        valid_types.append(tr.type_name)
                    elif tr.status == TypeStatus.RAG_MISS:
                        rag_misses.append(tr.type_name)
                    elif tr.status == TypeStatus.HALLUCINATION:
                        hallucinated.append(
                            f"{tr.type_name}"
                            + (f"→{tr.closest_real}" if tr.closest_real else "")
                        )

            gate_ok = not gate_report.had_hallucinations
            step_status = "CLEAN" if gate_ok else f"HALLUCINATION({len(hallucinated)})"
            step_overall.append(step_status)
            if gate_ok:
                clean_steps += 1

            print(f"  [gate] Summary: {summary}  |  status={step_status}", flush=True)
            if rag_misses:
                print(f"  [gate] RAG misses (added to CSV): {rag_misses}", flush=True)
            if hallucinated:
                print(f"  [gate] Hallucinations: {hallucinated}", flush=True)

            # ── Write to Excel ─────────────────────────────────────────────
            ws.cell(row=row, column=_col(step_n, "prompt"),
                    value=_clean(raw_prompt))
            ws.cell(row=row, column=_col(step_n, "raw_chain"),
                    value=_clean(raw_chain_str))
            ws.cell(row=row, column=_col(step_n, "gate_summary"),
                    value=_clean(summary))
            ws.cell(row=row, column=_col(step_n, "valid_types"),
                    value=_clean(", ".join(valid_types)))
            ws.cell(row=row, column=_col(step_n, "rag_miss_types"),
                    value=_clean(", ".join(rag_misses)))
            ws.cell(row=row, column=_col(step_n, "hallucinated_types"),
                    value=_clean(", ".join(hallucinated)))
            ws.cell(row=row, column=_col(step_n, "rewrites_done"),
                    value=str(rewrites_done))
            ws.cell(row=row, column=_col(step_n, "final_chain"),
                    value=_clean(final_chain_str))
            ws.cell(row=row, column=_col(step_n, "rewrite_feedback"),
                    value=_clean(gate_report.rewrite_feedback[:1000]))

        # Overall per case
        overall_str = (
            "ALL_CLEAN" if all(s == "CLEAN" for s in step_overall)
            else " | ".join(step_overall)
        )
        ws.cell(row=row, column=_overall_col(), value=overall_str)
        print(f"  [case overall] {overall_str}", flush=True)
        wb.save(result_file)

    print(f"\n{'='*60}", flush=True)
    print(f"Steps with no hallucination after rewrite: "
          f"{clean_steps}/{total_steps} = "
          f"{clean_steps/max(total_steps,1)*100:.1f}%",
          flush=True)
    print(f"Results → {result_file}", flush=True)


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(
        description="Extraction-only eval: causal chain + Hard Gate (no generation)"
    )
    p.add_argument("--openaiKey",    required=True)
    p.add_argument("--model",        default="gpt-4.1-mini")
    p.add_argument("--actionDBPath", default="src_1_reflector/action_db.xlsx")
    p.add_argument("--RAGApiPath",   default="RAGData/RAGAPIs_structured.csv",
                   help="Path to RAGAPIs_structured.csv (structured RAG)")
    p.add_argument("--resultPath",   default="result/")
    p.add_argument("--level",        type=int,  default=2)
    p.add_argument("--numCases",     type=int,  default=None)
    p.add_argument("--startCase",    type=int,  default=None,
                   help="0-indexed case to start from")
    p.add_argument("--cases",        type=int, nargs="+", default=None,
                   help="Specific 0-indexed case numbers (e.g. --cases 0 3 7)")
    p.add_argument("--runTag",       default="",
                   help="Optional suffix for result filename")
    p.add_argument("--sheet",        default=None,
                   help="Sheet name in the xlsx (default: active sheet)")
    args = p.parse_args()

    def _abs(s: str) -> str:
        return s if os.path.isabs(s) else os.path.join(_ROOT, s)

    Run(
        action_db_path = _abs(args.actionDBPath),
        result_path    = _abs(args.resultPath),
        rag_api_path   = _abs(args.RAGApiPath),
        api_key        = args.openaiKey,
        openai_model   = args.model,
        level          = args.level,
        num_cases      = args.numCases,
        start_case     = args.startCase,
        cases          = args.cases,
        run_tag        = args.runTag,
        sheet          = args.sheet,
    )


if __name__ == "__main__":
    main()
