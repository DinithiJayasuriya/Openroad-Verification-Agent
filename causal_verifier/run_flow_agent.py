"""run_flow_agent.py — Entry-point for the Level 2.1 flow-task pipeline.

For each prompt in the dataset:
  1. FlowChainExtractor  — identifies the tool, builds MultiPathCausalGraph
  2. FlowPipeline.run()  — generate → L4 verify → repair (up to --budget attempts)
  3. OpenROAD execution  — runs the final code in a live shell (if --openroadPath set)
  4. Excel output        — one row per case with all artefacts

Non-flow tasks (no tool keyword match) are recorded as SKIP so the sheet
stays aligned with the input dataset.

Usage:
  conda activate prompt
  python causal_verifier/run_flow_agent.py \\
      --dataset      path/to/tasks.xlsx \\
      --openaiKey    sk-... \\
      --openroadPath OpenROAD/build/src/openroad \\
      --runDir       src_1_reflector \\
      --resultPath   result/ \\
      --budget       3 \\
      --numCases     10

Excel output columns
--------------------
  A  prompt
  B  tool_id          (from graph, or "NOT_FLOW")
  C  constraint_prompt
  D  final_code
  E  l4_verdict       ("PASS" | "FAIL: C2 ..." | "SKIP")
  F  l4_issues        (JSON list of issue descriptions)
  G  attempts         (int — number of generate-verify cycles)
  H  openroad_result  ("PASS" | "FAIL" | "SKIPPED")
  I  openroad_output  (first 1000 chars of stdout)
"""

import argparse
import json
import os
import queue
import re
import sys
import threading
import time
from typing import Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

# ── path setup ────────────────────────────────────────────────────────────────
_ROOT       = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_CAUSAL_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(_ROOT, "src_1_reflector"))
sys.path.insert(0, os.path.join(_ROOT, "src_1_agentic"))
sys.path.insert(0, _CAUSAL_DIR)

from util import (
    runOpenROADShell, sendCommandOpenROAD, processCodeString,
    clearQueue, readOpenROADOutput,
)

from flow_pipeline import FlowPipeline


# ── helpers ───────────────────────────────────────────────────────────────────

def _clean(text: str) -> str:
    """Strip ANSI / control chars and illegal xlsx characters."""
    text = re.sub(r'\x1b\][^\x07]*\x07', '', str(text or ""))
    text = re.sub(r'\x1b\[[0-9;]*[a-zA-Z]', '', text)
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', text)
    return ILLEGAL_CHARACTERS_RE.sub("", text.strip())


# ── OpenROAD execution ────────────────────────────────────────────────────────

def execute_in_openroad(
    code: str,
    openroad_path: str,
    slave_fd: int,
    output_queue: queue.Queue,
    run_dir: str = "",
    load_design_time: int = 5,
    max_wait_time: int = 120,
    command_flush_time: float = 0.1,
) -> tuple:
    """Run *code* in a fresh OpenROAD shell. Returns (stdout, has_traceback)."""
    orig_dir = os.getcwd()
    openroad_path = os.path.abspath(openroad_path)
    if run_dir:
        os.chdir(run_dir)
    proc = None
    try:
        proc = runOpenROADShell(openroad_path, load_design_time, slave_fd, "")
        time.sleep(load_design_time)
        clearQueue(output_queue)

        cmd = processCodeString(code)
        while True:
            try:
                stdout, has_tb = sendCommandOpenROAD(
                    proc, cmd, output_queue, max_wait_time, command_flush_time,
                )
                break
            except RuntimeError:
                print("    [OpenROAD crashed] restarting...", flush=True)
                proc = runOpenROADShell(openroad_path, load_design_time, slave_fd, "")
                time.sleep(load_design_time)
                clearQueue(output_queue)

        proc.terminate()
        proc.wait()
        time.sleep(0.3)
        clearQueue(output_queue)
        return stdout, has_tb
    finally:
        os.chdir(orig_dir)


# ── Main run loop ─────────────────────────────────────────────────────────────

def Run(
    dataset_path:       str,
    result_path:        str,
    openai_key:         str,
    openai_model:       str           = "gpt-4.1-mini",
    openroad_path:      str           = "",
    run_dir:            str           = "",
    budget:             int           = 3,
    load_design_time:   int           = 5,
    max_wait_time:      int           = 120,
    command_flush_time: float         = 0.1,
    num_cases:          Optional[int] = None,
    start_case:         Optional[int] = None,
    run_tag:            str           = "",
):
    # ── load dataset ──────────────────────────────────────────────────────────
    ext = os.path.splitext(dataset_path)[1].lower()
    df  = pd.read_excel(dataset_path) if ext in (".xlsx", ".xls") else pd.read_csv(dataset_path)

    if "prompt" not in df.columns:
        print(f"ERROR: no 'prompt' column. Available: {list(df.columns)}")
        sys.exit(1)

    prompts = df["prompt"].dropna().tolist()
    if start_case:
        prompts = prompts[start_case:]
    if num_cases:
        prompts = prompts[:num_cases]

    print(f"[Init] {len(prompts)} prompt(s) loaded from {dataset_path}")

    # ── pipeline ──────────────────────────────────────────────────────────────
    pipeline = FlowPipeline(openai_key=openai_key, model=openai_model)
    print(f"[Init] FlowPipeline ready (model={openai_model}, budget={budget})")

    # ── OpenROAD PTY ──────────────────────────────────────────────────────────
    slave_fd     = None
    output_queue = None
    stop_event   = None
    if openroad_path:
        master_fd, slave_fd = os.openpty()
        output_queue        = queue.Queue()
        stop_event          = threading.Event()
        threading.Thread(
            target=readOpenROADOutput,
            args=(master_fd, output_queue, "STDOUT", stop_event),
            daemon=True,
        ).start()
        print("[Init] OpenROAD PTY ready.")
    else:
        print("[Init] --openroadPath not set — OpenROAD execution will be skipped.")

    # ── result workbook ───────────────────────────────────────────────────────
    os.makedirs(result_path, exist_ok=True)
    dataset_tag  = os.path.splitext(os.path.basename(dataset_path))[0]
    tag_suffix   = f"_{run_tag}" if run_tag else ""
    result_file  = os.path.join(
        result_path,
        f"{openai_model.replace('.', '-')}__flow_L21_{dataset_tag}{tag_suffix}.xlsx"
    )

    if os.path.exists(result_file):
        wb = load_workbook(result_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        headers = [
            "prompt", "tool_id", "constraint_prompt",
            "final_code", "l4_verdict", "l4_issues",
            "attempts", "openroad_result", "openroad_output",
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

    # ── case loop ─────────────────────────────────────────────────────────────
    passed_l4   = 0
    passed_ora  = 0
    total       = 0

    for case_i, task in enumerate(prompts):
        task   = str(task).strip()
        row    = (start_case or 0) + case_i + 2   # +2: header + 0-index
        total += 1

        print(f"\n{'='*60}", flush=True)
        print(f"[{case_i+1}/{len(prompts)}] {task}", flush=True)

        # ── run flow pipeline ─────────────────────────────────────────────────
        result = pipeline.run(task, budget=budget)

        # ── L4 verdict string ─────────────────────────────────────────────────
        if result.graph is None:
            l4_verdict  = "SKIP (not a flow task)"
            l4_issues   = "[]"
            tool_id     = "NOT_FLOW"
            attempts_n  = 0
        elif result.passed:
            l4_verdict  = "PASS"
            l4_issues   = "[]"
            tool_id     = result.graph.action_path.tool_def.tool_id
            attempts_n  = len(result.attempts)
            passed_l4  += 1
        else:
            issues      = result.l4_result.issues if result.l4_result else []
            l4_verdict  = "FAIL: " + "; ".join(i.check for i in issues)
            l4_issues   = json.dumps([i.description for i in issues])
            tool_id     = result.graph.action_path.tool_def.tool_id
            attempts_n  = len(result.attempts)

        print(f"  [L4] {l4_verdict}  (attempts={attempts_n})", flush=True)

        # ── OpenROAD execution ────────────────────────────────────────────────
        ora_result = "SKIPPED"
        ora_output = ""

        if openroad_path and result.final_code:
            print("  [OpenROAD] Running...", flush=True)
            try:
                stdout, has_tb = execute_in_openroad(
                    result.final_code, openroad_path, slave_fd, output_queue,
                    run_dir=run_dir,
                    load_design_time=load_design_time,
                    max_wait_time=max_wait_time,
                    command_flush_time=command_flush_time,
                )
                ora_result = "FAIL" if has_tb else "PASS"
                ora_output = stdout[:1000]
                if not has_tb:
                    passed_ora += 1
            except Exception as exc:
                ora_result = "FAIL"
                ora_output = f"[exec error] {exc}"

            print(f"  [OpenROAD] {ora_result}", flush=True)
            if ora_output:
                for ln in ora_output.splitlines()[:5]:
                    print(f"    {ln}", flush=True)

        # ── write row ─────────────────────────────────────────────────────────
        ws.cell(row=row, column=1, value=_clean(task))
        ws.cell(row=row, column=2, value=_clean(tool_id))
        ws.cell(row=row, column=3, value=_clean(result.constraint_prompt))
        ws.cell(row=row, column=4, value=_clean(result.final_code))
        ws.cell(row=row, column=5, value=_clean(l4_verdict))
        ws.cell(row=row, column=6, value=_clean(l4_issues))
        ws.cell(row=row, column=7, value=attempts_n)
        ws.cell(row=row, column=8, value=ora_result)
        ws.cell(row=row, column=9, value=_clean(ora_output))
        wb.save(result_file)

    # ── summary ───────────────────────────────────────────────────────────────
    print(f"\n{'='*60}", flush=True)
    print(f"L4 passed   : {passed_l4}/{total} = {passed_l4/max(total,1)*100:.1f}%")
    if openroad_path:
        print(f"OpenROAD    : {passed_ora}/{total} = {passed_ora/max(total,1)*100:.1f}%")
    print(f"Results     → {result_file}")

    if stop_event:
        stop_event.set()


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(
        description="Level 2.1 flow-task pipeline: generate → L4 verify → repair → execute"
    )
    p.add_argument("--dataset",          required=True,
                   help="Path to .xlsx or .csv with a 'prompt' column")
    p.add_argument("--openaiKey",        required=True)
    p.add_argument("--model",            default="gpt-4.1-mini")
    p.add_argument("--resultPath",       default="result/")
    p.add_argument("--openroadPath",     default="",
                   help="Path to openroad binary (enables execution)")
    p.add_argument("--runDir",           default="src_1_reflector",
                   help="Working dir for OpenROAD shell (loadDesign uses relative paths)")
    p.add_argument("--budget",           type=int, default=3,
                   help="Max generate-verify cycles per task")
    p.add_argument("--numCases",         type=int, default=None)
    p.add_argument("--startCase",        type=int, default=None,
                   help="0-indexed case to start from")
    p.add_argument("--loadDesignTime",   type=int, default=5)
    p.add_argument("--maxWaitTime",      type=int, default=120)
    p.add_argument("--commandFlushTime", type=float, default=0.1)
    p.add_argument("--runTag",           default="",
                   help="Optional tag appended to result filename")
    args = p.parse_args()

    def _abs(s: str) -> str:
        return s if (not s or os.path.isabs(s)) else os.path.join(_ROOT, s)

    Run(
        dataset_path       = _abs(args.dataset),
        result_path        = _abs(args.resultPath),
        openai_key         = args.openaiKey,
        openai_model       = args.model,
        openroad_path      = _abs(args.openroadPath) if args.openroadPath else "",
        run_dir            = _abs(args.runDir) if args.runDir else "",
        budget             = args.budget,
        load_design_time   = args.loadDesignTime,
        max_wait_time      = args.maxWaitTime,
        command_flush_time = args.commandFlushTime,
        num_cases          = args.numCases,
        start_case         = args.startCase,
        run_tag            = args.runTag,
    )


if __name__ == "__main__":
    main()
