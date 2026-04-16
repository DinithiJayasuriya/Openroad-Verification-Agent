"""run_causal_agent_sequential_option3.py — Option 3: full chain build then single codegen.

Option 3 behaviour
------------------
  Step 1  : Extract causal chain fresh from sub-prompt 1.  No code generation yet.
  Step N>1: Read sub-prompt N and EXTEND the accumulated chain with the new nodes
            needed (same extension logic as Option 2).  Still no code generation.
  After all sub-prompts: generate ONE combined code script from the fully extended
            chain and the complete task description, run verifier + controller loop,
            execute ONCE in OpenROAD.

Key difference vs Option 2
  Option 2 generates + executes code after every sub-prompt (gated on prior pass).
  Option 3 generates + executes code only ONCE at the very end.

Controlled by --level:
  level 1 — uses only step 1 (level_one column)
  level 2 — step 1 + step 2 (level_one + level_two columns)
  level 3 — all 3 sub-prompts (level_one + level_two + level_three columns)

Usage:
  python causal_verifier/run_causal_agent_sequential_option3.py \\
    --openaiKey sk-... \\
    --actionDBPath src_1_reflector/action_db.xlsx \\
    --RAGApiPath RAGData/RAGAPIs.csv \\
    --RAGCodePiecePath RAGData/RAGCodePiece.csv \\
    --openroadPath OpenROAD/build/src/openroad \\
    --runDir src_1_reflector \\
    --resultPath result/ \\
    --level 3 --budget 6 --numCases 5
"""

import argparse
import os
import queue
import re
import sys
import threading
import time
from typing import List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from sentence_transformers import SentenceTransformer

# ── path setup ────────────────────────────────────────────────────────────────
_ROOT       = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_CAUSAL_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(_ROOT, "src_1_reflector"))
sys.path.insert(0, os.path.join(_ROOT, "src_1_agentic"))
sys.path.insert(0, _CAUSAL_DIR)

from util import (
    runOpenROADShell, sendCommandOpenROAD, processCodeString,
    clearQueue, readOpenROADOutput,
)
from verifier import OpenROADStaticVerifier

from causal_state    import CausalAgentState
from causal_verifier import CausalVerifier
from llm_verifier    import CausalLLMVerifier

# Re-use helpers from run_causal_agent
from run_causal_agent import (
    _clean_output,
    _call_openai, _parse_multi_path_chain, _rag_query_for_edge,
    _rewrite_traversal_chain, bootstrap_causal_rag,
    bootstrap_generate, bootstrap_static_verify,
    run_bootstrap,
    CHAIN_SYSTEM_PROMPT,
)
from controller  import CausalController
from dispatcher  import CausalDispatcher


# ─────────────────────────────────────────────────────────────────────────────
# Option 2 — chain extension helpers
# ─────────────────────────────────────────────────────────────────────────────

_EXTEND_SYSTEM_PROMPT = """\
You are an OpenROAD Python API expert.

You have an existing causal acquisition chain from a prior step.
A new sub-task requires additional objects. Your job is to EXTEND the chain by
adding only the new nodes needed for the new sub-task.

RULES:
- The existing chain nodes are already acquired in the live OpenROAD session — do NOT repeat them.
- Only add NEW terminal nodes not already present in the existing chain.
- Each new path starts from the closest already-acquired node (do not restart from openroad.Design if a shorter path exists through an already-acquired node).
- Use exact OpenROAD Python type names.
- If the new sub-task needs no additional objects beyond what the existing chain provides, output: NO_EXTENSION

OUTPUT FORMAT (no markdown, no JSON):
Existing chain: <shown to you below — do not repeat>
New Path 1 (<NewTerminalType>): <AlreadyAcquiredNode> -> ... -> <NewTerminalType>
New Path 2 (<NewTerminalType>): <AlreadyAcquiredNode> -> ... -> <NewTerminalType>
Action Node: <action>(<TerminalType1>, ...)
"""


def bootstrap_causal_extend(
    state: CausalAgentState,
    prev_state: CausalAgentState,
    api_key: str,
    model: str,
    embed_model,
    metadata: list,
    embeddings,
) -> None:
    """Option 2 bootstrap step 1: extend the previous step's chain for a new sub-task.

    Reuses prev_state.chain and prev_state.edge_apis for already-acquired nodes.
    Only runs RAG for genuinely new edges.  Updates state.chain / state.all_edges /
    state.edge_apis in place.
    """
    prev_chain_str = " -> ".join(prev_state.chain) if prev_state.chain else "(empty)"

    user_msg = (
        f"Existing chain from previous step: {prev_chain_str}\n\n"
        f"New sub-task: {state.task}"
    )

    text = _call_openai(
        messages=[
            {"role": "system", "content": _EXTEND_SYSTEM_PROMPT},
            {"role": "user",   "content": user_msg},
        ],
        api_key=api_key, model=model, max_tokens=300,
    )

    print(f"  [extend/raw] {text[:200]}", flush=True)

    # If the LLM says no new nodes are needed, inherit the previous chain as-is
    if not text.strip() or "NO_EXTENSION" in text.upper():
        print("  [extend] NO_EXTENSION — inheriting previous chain unchanged", flush=True)
        state.chain     = list(prev_state.chain)
        state.paths     = list(prev_state.paths)
        state.all_edges = list(prev_state.all_edges)
        state.edge_apis = list(prev_state.edge_apis)
        state.api_summary = prev_state.api_summary
        state.add_bootstrap_obs("causal_extend", "NO_EXTENSION — reused previous chain")
        return

    # Parse new paths from the extension output
    # Re-use the existing parser but prefix lines to match its "Path N (...): ..." pattern
    normalised = re.sub(r'(?i)^New\s+(Path)', r'\1', text, flags=re.MULTILINE)
    new_paths, action_node = _parse_multi_path_chain(normalised)

    if not new_paths:
        # Fallback: re-extract fresh (LLM output unparseable)
        print("  [extend] parse failed — falling back to fresh extraction", flush=True)
        from run_causal_agent import bootstrap_causal_extract
        bootstrap_causal_extract(state, api_key, model)
        bootstrap_causal_rag(state, embed_model, metadata, embeddings)
        return

    # --- Merge: existing chain nodes + new extension nodes ---
    existing_nodes: List[str] = list(prev_state.chain)
    existing_set   = set(existing_nodes)

    # Collect only the genuinely new nodes (preserve order, deduplicate)
    new_nodes: List[str] = []
    for path in new_paths:
        for node in path:
            if node not in existing_set and node not in new_nodes:
                new_nodes.append(node)

    # Build merged chain: existing nodes first, then new terminal nodes appended
    merged_chain = existing_nodes + new_nodes
    state.chain      = merged_chain
    state.action_node = action_node

    # Build all_edges: reuse previous edges + add edges for each new path
    existing_edges_set = {(s, t) for s, t in prev_state.all_edges}
    merged_edges: List[Tuple[str, str]] = list(prev_state.all_edges)

    for path in new_paths:
        for src, tgt in zip(path[:-1], path[1:]):
            if (src, tgt) not in existing_edges_set:
                merged_edges.append((src, tgt))
                existing_edges_set.add((src, tgt))

    state.all_edges = merged_edges
    state.paths     = (prev_state.paths or [list(prev_state.chain)]) + new_paths

    # Apply traversal-pattern guard to merged chain
    state.chain     = _rewrite_traversal_chain(state.task, state.chain)

    # --- RAG: reuse existing edge_apis, run retrieval only for new edges ---
    prev_edge_map: dict = {}
    for edge, api in zip(prev_state.all_edges, prev_state.edge_apis or []):
        prev_edge_map[edge] = api

    edge_apis  = []
    api_parts  = []
    for edge in state.all_edges:
        if edge in prev_edge_map:
            hit = prev_edge_map[edge]
            edge_apis.append(hit)
            if hit:
                method = hit["function_name"].split("(")[0].split(".")[-1].strip()
                api_parts.append(
                    f"{edge[0].split('.')[-1]}->{edge[1].split('.')[-1]}: {method} [reused]"
                )
                print(f"  [extend/rag] [{edge[0]}] -> [{edge[1]}]  reused: {hit['function_name']}",
                      flush=True)
            else:
                api_parts.append(
                    f"{edge[0].split('.')[-1]}->{edge[1].split('.')[-1]}: NOT FOUND [reused]"
                )
        else:
            hit = _rag_query_for_edge(edge[0], edge[1], embed_model, metadata, embeddings)
            edge_apis.append(hit)
            if hit:
                method = hit["function_name"].split("(")[0].split(".")[-1].strip()
                api_parts.append(
                    f"{edge[0].split('.')[-1]}->{edge[1].split('.')[-1]}: {method} [new]"
                )
                print(f"  [extend/rag] [{edge[0]}] -> [{edge[1]}]  NEW hit: {hit['function_name']}",
                      flush=True)
            else:
                api_parts.append(
                    f"{edge[0].split('.')[-1]}->{edge[1].split('.')[-1]}: NOT FOUND [new]"
                )
                print(f"  [extend/rag] [{edge[0]}] -> [{edge[1]}]  NEW hit: [none]", flush=True)

    state.edge_apis   = edge_apis
    state.api_summary = " | ".join(api_parts)
    state.add_bootstrap_obs(
        "causal_extend",
        f"extended chain: {' -> '.join(state.chain)}  |  edges: {state.api_summary}",
    )


# ─────────────────────────────────────────────────────────────────────────────
# Option 2 case runner — respects state._bootstrap_done
# ─────────────────────────────────────────────────────────────────────────────

def run_case_option3(
    task: str,
    state: CausalAgentState,
    api_key: str,
    model: str,
    embed_model,
    metadata: list,
    embeddings,
    static_verifier: OpenROADStaticVerifier,
    causal_verifier: CausalVerifier,
    code_pieces: Optional[list] = None,
    llm_verifier=None,
) -> CausalAgentState:
    """Bootstrap → controller loop → commit.

    Identical to run_case except: when state._bootstrap_done is True (chain was
    already extended by bootstrap_causal_extend), skip extract+RAG and jump
    straight to generate+verify.
    """
    print(f"\n{'─'*60}", flush=True)
    print(f"TASK: {task}", flush=True)
    print(f"BUDGET: {state.max_budget} controller actions", flush=True)

    if getattr(state, "_bootstrap_done", False):
        # Chain + edge_apis already set by bootstrap_causal_extend — skip extract+RAG
        print(f"[Pipeline] [extended] generate → causal_verify (chain reused)", flush=True)
        bootstrap_generate(state, api_key, model)
        bootstrap_static_verify(state, static_verifier, causal_verifier)
        print(f"[Bootstrap] done — committing generated code for OpenROAD execution\n",
              flush=True)
    else:
        run_bootstrap(state, api_key, model, embed_model, metadata, embeddings,
                      static_verifier, causal_verifier)

    # ── same commit / controller logic as run_case ─────────────────────────────
    if state.static_result and state.static_result.passed:
        if llm_verifier is not None and state.current_code:
            from dispatcher import _build_chain_context
            chain_ctx = _build_chain_context(state.chain, state.edge_apis)
            lv_snap   = llm_verifier.verify(task, state.current_code, chain_ctx)
            state.llm_result = lv_snap
            if lv_snap.passed:
                print(f"  [loop] bootstrap PASS + LLM PASS (conf={lv_snap.confidence:.2f})"
                      f" — committing directly", flush=True)
                state.committed_code = state.best_code or state.current_code
                state.committed      = True
                return state
            else:
                print(f"  [loop] bootstrap static PASS but LLM FAIL(L5) "
                      f"conf={lv_snap.confidence:.2f} — entering controller loop",
                      flush=True)
        else:
            print(f"  [loop] bootstrap passed — committing directly", flush=True)
            state.committed_code = state.best_code or state.current_code
            state.committed      = True
            return state

    ctrl       = CausalController(api_key=api_key, model=model)
    dispatcher = CausalDispatcher(
        api_key=api_key, model=model,
        embed_model=embed_model, metadata=metadata, embeddings=embeddings,
        static_verifier=static_verifier, causal_verifier=causal_verifier,
        code_pieces=code_pieces or [],
        llm_verifier=llm_verifier,
    )
    dispatcher.reset_conversation(state)

    while not state.committed and state.budget_remaining > 0:
        action      = ctrl.decide(state)
        observation = dispatcher.execute(action, state)
        state.add_observation(action.next_action, observation, action.diagnosis or "")

    if not state.committed:
        state.committed_code = state.best_code or state.current_code
        state.committed      = True
        print(f"  [loop] budget exhausted — committing best (score={state.best_score:.3f})",
              flush=True)

    return state


# ─────────────────────────────────────────────────────────────────────────────
# Case loader  (identical logic to run_agent_sequential.py)
# ─────────────────────────────────────────────────────────────────────────────

def load_cases(path: str, level: int) -> List[Tuple[str, List[str]]]:
    """Return list of (complex_prompt, steps[]) from a Format-B dataset.

    Expected column layout (one row per case):
        complex_prompt | step_1 | step_2 | step_3 | ... | step_N

    level=N reads columns 1..N (step_1 through step_N).
    Rows with an empty complex_prompt are skipped.
    Steps beyond available columns are returned as empty strings.
    """
    import openpyxl
    wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    cases = []
    for row in rows[1:]:                          # skip header row
        if not row[0] or not str(row[0]).strip():
            continue
        complex_prompt = str(row[0]).strip()
        steps = []
        for col_idx in range(1, level + 1):       # cols 1..level
            val = row[col_idx] if len(row) > col_idx else None
            steps.append(str(val).strip() if val and str(val).strip() else "")
        cases.append((complex_prompt, steps))
    return cases


# ─────────────────────────────────────────────────────────────────────────────
# OpenROAD execution helper (runs in an already-open process)
# ─────────────────────────────────────────────────────────────────────────────

def _send_code(proc, code: str, oq: queue.Queue,
               max_wait: int, flush_time: float) -> Tuple[str, bool]:
    cmd = processCodeString(code)
    return sendCommandOpenROAD(proc, cmd, oq, max_wait, flush_time)


# ─────────────────────────────────────────────────────────────────────────────
# Main eval loop
# ─────────────────────────────────────────────────────────────────────────────

def Run(
    action_db_path:     str,
    result_path:        str,
    rag_api_path:       str,
    rag_code_path:      str,
    openroad_path:      str,
    api_key:            str,
    openai_model:       str           = "gpt-4.1-mini",
    level:              int           = 2,
    budget:             int           = 6,
    load_design_time:   int           = 5,
    max_wait_time:      int           = 120,
    command_flush_time: float         = 0.1,
    num_cases:          Optional[int] = None,
    start_case:         Optional[int] = None,
    cases:              Optional[List[int]] = None,
    run_dir:            str           = "",
    run_tag:            str           = "",
):
    # ── RAG setup ─────────────────────────────────────────────────────────────
    print("\nLoading embedding model...", flush=True)
    embed_model = SentenceTransformer("all-MiniLM-L6-v2")

    rag_df    = pd.read_csv(rag_api_path)
    metadata  = []
    documents = []
    for _, row in rag_df.iterrows():
        desc = str(row.get("Description:", "")).strip()
        if not desc or desc.lower() == "nan":
            continue
        documents.append(f"OpenROAD Python API Description:{desc}")
        metadata.append(row.to_dict())
    print(f"  [RAG] Encoding {len(documents)} API entries...", flush=True)
    embeddings = embed_model.encode(documents, convert_to_tensor=True,
                                    show_progress_bar=False)

    # ── verifiers ─────────────────────────────────────────────────────────────
    static_ver = OpenROADStaticVerifier(rag_api_path)
    causal_ver = CausalVerifier(metadata=metadata)
    llm_ver    = CausalLLMVerifier(api_key=api_key, model=openai_model, fail_open=True)
    print("[Init] Static + causal + LLM semantic verifiers ready.", flush=True)

    # ── code examples (RAGCodePiece.csv) ──────────────────────────────────────
    code_pieces = []
    if rag_code_path and os.path.isfile(rag_code_path):
        cp_df = pd.read_csv(rag_code_path)
        for _, row in cp_df.iterrows():
            desc = str(row.get("Description:", "")).strip()
            code = str(row.get("Code Piece:", "")).strip()
            if desc and code and desc.lower() != "nan" and code.lower() != "nan":
                code_pieces.append({"description": desc, "code": code})
        print(f"[Init] Loaded {len(code_pieces)} code examples.", flush=True)
    else:
        print("[Init] RAGCodePiecePath not set — code examples disabled.", flush=True)

    # ── OpenROAD PTY ──────────────────────────────────────────────────────────
    master_fd, slave_fd = os.openpty()
    oq                  = queue.Queue()
    threading.Thread(
        target=readOpenROADOutput,
        args=(master_fd, oq, "STDOUT", threading.Event()),
        daemon=True,
    ).start()
    print("[Init] OpenROAD PTY ready.", flush=True)

    # ── load + slice cases ────────────────────────────────────────────────────
    all_cases = load_cases(action_db_path, level)
    if cases is not None:
        indexed_cases = [(i, all_cases[i][0], all_cases[i][1])
                         for i in cases if i < len(all_cases)]
    else:
        if start_case is not None:
            all_cases = all_cases[start_case:]
        if num_cases is not None:
            all_cases = all_cases[:num_cases]
        offset = start_case or 0
        indexed_cases = [(offset + i, cp, steps)
                         for i, (cp, steps) in enumerate(all_cases)]
    print(f"[Init] Level={level}, running {len(indexed_cases)} case(s).", flush=True)

    # ── result workbook ───────────────────────────────────────────────────────
    os.makedirs(result_path, exist_ok=True)
    dataset_tag  = os.path.splitext(os.path.basename(action_db_path))[0]
    tag_suffix   = f"_{run_tag}" if run_tag else ""
    _result_file = os.path.join(
        result_path,
        f"{openai_model.replace('.', '-')}__causal_L{level}_{dataset_tag}_opt3{tag_suffix}.xlsx"
    )

    if os.path.exists(_result_file):
        wb = load_workbook(_result_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        # Option 3: one row per case — flat columns (no per-step code/result)
        _HEADERS = [
            "complex_prompt",
            "sub_prompts",          # all sub-prompt texts joined
            "final_chain",          # fully extended chain after all steps
            "chain_per_step",       # chain state after each extension step
            "node_apis",            # API summary for the final chain
            "generated_code",       # single code generated from full chain
            "causal_verdict",       # static+causal verifier verdict
            "openroad_result",      # PASS / FAIL
            "openroad_output",      # truncated stdout
            "budget_used",
            "lessons",
        ]
        for col, h in enumerate(_HEADERS, 1):
            ws.cell(row=1, column=col, value=h)

    def _clean(s):
        return ILLEGAL_CHARACTERS_RE.sub("", str(s or ""))

    passed_all = 0
    total      = 0

    # ── case loop ─────────────────────────────────────────────────────────────
    for case_i, complex_prompt, steps in indexed_cases:
        row   = case_i + 2
        total += 1
        print(f"\n{'='*60}", flush=True)
        print(f"CASE {case_i+1}/{len(indexed_cases)}  (level={level}, {len(steps)} steps)",
              flush=True)

        # ── Phase 1: build the full chain by iterating over all sub-prompts ──
        # No code generation, no OpenROAD execution here.
        chain_states: List[CausalAgentState] = []   # one per sub-prompt
        chain_snapshots: List[str]           = []   # human-readable chain after each step

        for step_idx, raw_prompt in enumerate(steps):
            step_n = step_idx + 1
            if not raw_prompt:
                continue
            print(f"\n  ── [Chain build] Step {step_n}/{len(steps)}: {raw_prompt[:100]}",
                  flush=True)

            # Create a lightweight state for chain extraction only
            chain_state = CausalAgentState(task=raw_prompt, max_budget=0)

            if step_idx == 0 or not chain_states:
                # Step 1: fresh causal extraction + RAG
                print(f"  [Option3] Step 1 — fresh chain extraction", flush=True)
                from run_causal_agent import bootstrap_causal_extract
                bootstrap_causal_extract(chain_state, api_key, openai_model)
                if chain_state.chain:
                    from run_causal_agent import _rewrite_traversal_chain
                    chain_state.chain = _rewrite_traversal_chain(
                        chain_state.task, chain_state.chain
                    )
                bootstrap_causal_rag(
                    chain_state, embed_model, metadata, embeddings
                )
            else:
                # Step N>1: extend the accumulated chain
                prev_chain_state = next(
                    (s for s in reversed(chain_states) if s is not None), None
                )
                if prev_chain_state is not None:
                    print(
                        f"  [Option3] Step {step_n} — extending chain: "
                        f"{' -> '.join(prev_chain_state.chain)}",
                        flush=True,
                    )
                    bootstrap_causal_extend(
                        chain_state, prev_chain_state,
                        api_key=api_key, model=openai_model,
                        embed_model=embed_model, metadata=metadata,
                        embeddings=embeddings,
                    )
                else:
                    print(f"  [Option3] Step {step_n} — fallback: fresh extraction",
                          flush=True)
                    from run_causal_agent import bootstrap_causal_extract
                    bootstrap_causal_extract(chain_state, api_key, openai_model)
                    bootstrap_causal_rag(chain_state, embed_model, metadata, embeddings)

            chain_states.append(chain_state)
            chain_snapshots.append(" -> ".join(chain_state.chain))
            print(f"  [Option3] Chain after step {step_n}: {chain_snapshots[-1]}",
                  flush=True)

        if not chain_states:
            print(f"  [Case {case_i+1}] No valid steps — skipping", flush=True)
            total -= 1
            continue

        # ── Phase 2: generate ONE code from the fully extended chain ──────────
        final_chain_state = chain_states[-1]

        # Task for generation = complex_prompt + all sub-prompts as context
        sub_prompt_block = "\n".join(
            f"  Sub-task {i+1}: {steps[i]}"
            for i in range(len(chain_states))
            if steps[i]
        )
        combined_task = (
            f"{complex_prompt}\n\n"
            f"This task consists of the following sub-tasks that must ALL be "
            f"accomplished in a single script:\n{sub_prompt_block}"
        )

        # Build the final generation state with the merged chain
        gen_state = CausalAgentState(task=combined_task, max_budget=budget)
        gen_state.chain       = list(final_chain_state.chain)
        gen_state.paths       = list(final_chain_state.paths or [])
        gen_state.all_edges   = list(final_chain_state.all_edges or [])
        gen_state.edge_apis   = list(final_chain_state.edge_apis or [])
        gen_state.api_summary = final_chain_state.api_summary
        gen_state.action_node = final_chain_state.action_node
        gen_state._bootstrap_done = True   # chain+RAG already complete

        print(f"\n  [Option3] Final chain: {' -> '.join(gen_state.chain)}", flush=True)
        print(f"  [Option3] Generating whole code...", flush=True)

        # Fresh OpenROAD process for this case (single execution)
        orig_dir = os.getcwd()
        if run_dir:
            os.chdir(run_dir)
        proc = runOpenROADShell(openroad_path, load_design_time, slave_fd, "")
        if run_dir:
            os.chdir(orig_dir)
        time.sleep(load_design_time)
        clearQueue(oq)

        gen_state = run_case_option3(
            combined_task, gen_state,
            api_key=api_key, model=openai_model,
            embed_model=embed_model, metadata=metadata, embeddings=embeddings,
            static_verifier=static_ver, causal_verifier=causal_ver,
            code_pieces=code_pieces,
            llm_verifier=llm_ver,
        )

        # causal verifier verdict
        sv = gen_state.static_result
        causal_verdict = (
            "PASS" if (sv and sv.passed)
            else (f"FAIL(L{sv.layer_failed}): {'; '.join(sv.issues[:2])}" if sv else "N/A")
        )

        # ── Phase 3: execute once in OpenROAD ────────────────────────────────
        stdout, has_tb = "", True
        if gen_state.committed_code:
            try:
                stdout, has_tb = _send_code(
                    proc, gen_state.committed_code, oq,
                    max_wait_time, command_flush_time,
                )
            except Exception as exc:
                stdout   = f"[OpenROAD error] {exc}"
                has_tb   = True

        try:
            proc.terminate()
            proc.wait()
        except Exception:
            pass

        passed = bool(gen_state.committed_code) and not has_tb
        if passed:
            passed_all += 1

        overall = "PASS" if passed else "FAIL"
        print(f"  [Overall] {overall}", flush=True)
        if not passed and stdout:
            print(f"  [error]\n{stdout.strip()}", flush=True)

        # ── write Excel ───────────────────────────────────────────────────────
        ws.cell(row=row, column=1,  value=_clean(complex_prompt))
        ws.cell(row=row, column=2,  value=_clean(" | ".join(s for s in steps if s)))
        ws.cell(row=row, column=3,  value=_clean(" -> ".join(gen_state.chain)))
        ws.cell(row=row, column=4,  value=_clean(" || ".join(chain_snapshots)))
        ws.cell(row=row, column=5,  value=_clean(gen_state.api_summary))
        ws.cell(row=row, column=6,  value=_clean(gen_state.committed_code))
        ws.cell(row=row, column=7,  value=causal_verdict)
        ws.cell(row=row, column=8,  value=overall)
        ws.cell(row=row, column=9,  value=_clean(stdout[:2000]))
        ws.cell(row=row, column=10, value=gen_state.max_budget - gen_state.budget_remaining)
        ws.cell(row=row, column=11, value=_clean(" | ".join(gen_state.lessons)))
        wb.save(_result_file)

    print(f"\n{'='*60}", flush=True)
    print(f"Passed: {passed_all}/{total} = "
          f"{passed_all/max(total,1)*100:.1f}%", flush=True)
    print(f"Results → {_result_file}", flush=True)


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(
        description="Option 3: build full chain across all sub-prompts, generate whole code once"
    )
    p.add_argument("--openaiKey",        required=True)
    p.add_argument("--model",            default="gpt-4.1-mini")
    p.add_argument("--actionDBPath",     default="src_1_reflector/action_db.xlsx")
    p.add_argument("--resultPath",       default="result/")
    p.add_argument("--RAGApiPath",       default="RAGData/RAGAPIs.csv")
    p.add_argument("--RAGCodePiecePath", default="RAGData/RAGCodePiece.csv")
    p.add_argument("--openroadPath",     default="OpenROAD/build/src/openroad")
    p.add_argument("--runDir",           default="src_1_reflector",
                   help="Working dir for OpenROAD shell")
    p.add_argument("--level",            type=int, default=2,
                   help="Number of steps to run (1..N)")
    p.add_argument("--budget",           type=int, default=6,
                   help="Causal controller action budget per step")
    p.add_argument("--loadDesignTime",   type=int, default=5)
    p.add_argument("--maxWaitTime",      type=int, default=120)
    p.add_argument("--commandFlushTime", type=float, default=0.1)
    p.add_argument("--numCases",         type=int, default=None)
    p.add_argument("--startCase",        type=int, default=None,
                   help="0-indexed case to start from")
    p.add_argument("--cases",            type=int, nargs="+", default=None,
                   help="Specific 0-indexed case numbers to run (e.g. --cases 2 6)")
    p.add_argument("--runTag",           default="",
                   help="Optional tag appended to result filename to avoid overwriting")
    args = p.parse_args()

    def _abs(s):
        return s if os.path.isabs(s) else os.path.join(_ROOT, s)

    Run(
        action_db_path     = _abs(args.actionDBPath),
        result_path        = _abs(args.resultPath),
        rag_api_path       = _abs(args.RAGApiPath),
        rag_code_path      = _abs(args.RAGCodePiecePath),
        openroad_path      = _abs(args.openroadPath),
        api_key            = args.openaiKey,
        openai_model       = args.model,
        level              = args.level,
        budget             = args.budget,
        load_design_time   = args.loadDesignTime,
        max_wait_time      = args.maxWaitTime,
        command_flush_time = args.commandFlushTime,
        num_cases          = args.numCases,
        start_case         = args.startCase,
        cases              = args.cases,
        run_dir            = _abs(args.runDir) if args.runDir else "",
        run_tag            = args.runTag,
    )


if __name__ == "__main__":
    main()
