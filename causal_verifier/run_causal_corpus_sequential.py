"""run_causal_corpus_sequential.py — Causal agent runner for EDA-Corpus-v2/TestSet.xlsx.

Drop-in companion to run_causal_agent_sequential.py that reads the two-sheet
EDA-Corpus-v2 format (Sheet 'Prompt' + Sheet 'Code') instead of action_db.xlsx.

All causal pipeline logic (chain extraction, static + causal + LLM verifiers,
controller loop) is identical to run_causal_agent_sequential.py.

Ablation flags
--------------
  --noLLMVerifier      Disable LLM (GPT-4o) semantic verifier layer
  --noCausalVerifier   Disable causal chain verifier (implies --noLLMVerifier)

These let you run all three ablation conditions from one script:
  Run A:  (default)              static + causal + LLM  →  Full Pipeline
  Run B:  --noLLMVerifier        static + causal only
  Run C:  --noCausalVerifier     static only

Logging
-------
  Each result row includes:
    latency_s       — wall-clock seconds for the full case (LLM + OpenROAD)
    openroad_calls  — number of times OpenROAD was actually executed (always 1
                      for our pipeline; baseline would be budget_used)
    budget_used     — number of controller LLM iterations consumed

Usage:
  conda activate prompt
  python causal_verifier/run_causal_corpus_sequential.py \\
      --testSetPath    EDA-Corpus-v2/TestSet.xlsx \\
      --RAGApiPath     RAGData/RAGAPIs.csv \\
      --RAGCodePiecePath RAGData/RAGCodePiece.csv \\
      --openaiKey      sk-... \\
      --openroadPath   OpenROAD/build/src/openroad \\
      --runDir         src_1_reflector \\
      --numCases       20 \\
      --budget         6 \\
      --resultPath     result/causal_corpus_full.xlsx

  # Ablation: static + causal only (no LLM verifier)
  python causal_verifier/run_causal_corpus_sequential.py ... --noLLMVerifier \\
      --resultPath result/causal_corpus_no_llm.xlsx

  # Ablation: static only (no causal or LLM verifier)
  python causal_verifier/run_causal_corpus_sequential.py ... --noCausalVerifier \\
      --resultPath result/causal_corpus_static_only.xlsx
"""

import argparse
import os
import queue
import re
import sys
import threading
import time
from typing import List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from sentence_transformers import SentenceTransformer

# ── path setup ────────────────────────────────────────────────────────────────
_ROOT       = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_CAUSAL_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(_ROOT, "src_1_reflector"))
sys.path.insert(0, os.path.join(_ROOT, "src_1_agentic"))
sys.path.insert(0, _CAUSAL_DIR)

from util import (
    runOpenROADShell, sendCommandOpenROAD, processCodeString,
    clearQueue, readOpenROADOutput,
)
from verifier import OpenROADStaticVerifier, VerifierResult

from causal_state    import CausalAgentState, VerifierSnapshot
from causal_verifier import CausalVerifier
from llm_verifier    import CausalLLMVerifier

# Re-use run_case and helpers from run_causal_agent
from run_causal_agent import run_case, _clean_output


# ─────────────────────────────────────────────────────────────────────────────
# No-op verifier wrappers for ablation
# These satisfy the type interface expected by run_case / run_bootstrap without
# modifying any shared code.
# ─────────────────────────────────────────────────────────────────────────────

class _PassthroughStaticVerifier:
    """Always returns PASS — used when --noCausalVerifier disables static checks."""
    def verify(self, task: str, code: str, rag_context: str = "") -> VerifierResult:
        r = VerifierResult(passed=True, layer_failed=0, issues=[], feedback="")
        r.confidence   = 1.0
        r.warnings     = []
        r.is_soft_fail = False
        r.api_diffs    = []
        r.next_step    = "regenerate"
        return r


class _PassthroughCausalVerifier:
    """Always returns PASS — used when --noCausalVerifier is set."""
    def verify(self, code: str, chain: list, edge_apis: list) -> VerifierSnapshot:
        return VerifierSnapshot(
            passed=True, layer_failed=0, issues=[], feedback="", confidence=1.0
        )


# ─────────────────────────────────────────────────────────────────────────────
# Dataset loader  (EDA-Corpus-v2 two-sheet format)
# ─────────────────────────────────────────────────────────────────────────────

def load_corpus(path: str, num_cases: int = 0, prompt_file: str = "") -> Tuple[List[str], List[str]]:
    """Load prompts and ground-truth code from EDA-Corpus-v2 TestSet.xlsx.

    Expects:
      Sheet 'Prompt' : one prompt per row, no header (header=None)
      Sheet 'Code'   : one ground-truth script per row, same indexing

    Returns (prompts, gt_codes) as plain lists of strings.
    Blank / whitespace-only rows are skipped.

    If prompt_file is given, only prompts whose text matches a line in that
    file (stripped, exact substring match) are included.
    """
    prompt_df = pd.read_excel(path, "Prompt", header=None).rename(columns={0: "text"})
    code_df   = pd.read_excel(path, "Code",   header=None).rename(columns={0: "text"})

    filter_set = []
    if prompt_file and os.path.isfile(prompt_file):
        with open(prompt_file) as f:
            filter_set = [line.strip() for line in f if line.strip()]
        print(f"[Filter] Restricting to {len(filter_set)} prompts from {prompt_file}")

    prompts  = []
    gt_codes = []
    for p, c in zip(prompt_df["text"], code_df["text"]):
        p_str = str(p).strip()
        c_str = str(c).strip()
        if not p_str or p_str.lower() == "nan":
            continue
        if filter_set and not any(f in p_str or p_str in f for f in filter_set):
            continue
        prompts.append(p_str)
        gt_codes.append(c_str if c_str.lower() != "nan" else "")

    if num_cases > 0:
        prompts  = prompts[:num_cases]
        gt_codes = gt_codes[:num_cases]

    return prompts, gt_codes


# ─────────────────────────────────────────────────────────────────────────────
# OpenROAD execution helper
# ─────────────────────────────────────────────────────────────────────────────

def _send_code(proc, code: str, oq: queue.Queue,
               max_wait: int, flush_time: float) -> Tuple[str, bool]:
    cmd = processCodeString(code)
    return sendCommandOpenROAD(proc, cmd, oq, max_wait, flush_time)


# ─────────────────────────────────────────────────────────────────────────────
# Main eval loop
# ─────────────────────────────────────────────────────────────────────────────

def Run(
    test_set_path:       str,
    result_path:         str,
    rag_api_path:        str,
    rag_code_path:       str,
    openroad_path:       str,
    api_key:             str,
    openai_model:        str           = "gpt-4.1-mini",
    budget:              int           = 6,
    load_design_time:    int           = 5,
    max_wait_time:       int           = 120,
    command_flush_time:  float         = 0.1,
    num_cases:           Optional[int] = None,
    start_case:          Optional[int] = None,
    cases:               Optional[List[int]] = None,
    run_dir:             str           = "",
    run_tag:             str           = "",
    prompt_file:         str           = "",
    no_llm_verifier:     bool          = False,
    no_causal_verifier:  bool          = False,
):
    # ── resolve ablation mode ──────────────────────────────────────────────────
    # --noCausalVerifier implies --noLLMVerifier (nothing left above static)
    if no_causal_verifier:
        no_llm_verifier = True

    ablation_tag = (
        "static_only"    if no_causal_verifier else
        "no_llm_ver"     if no_llm_verifier    else
        "full_pipeline"
    )
    print(f"\n[Ablation mode] {ablation_tag}", flush=True)

    # ── RAG setup ─────────────────────────────────────────────────────────────
    print("Loading embedding model...", flush=True)
    embed_model = SentenceTransformer("all-MiniLM-L6-v2")

    rag_df    = pd.read_csv(rag_api_path)
    metadata  = []
    documents = []
    for _, row in rag_df.iterrows():
        desc = str(row.get("Description:", "")).strip()
        if not desc or desc.lower() == "nan":
            continue
        documents.append(f"OpenROAD Python API Description:{desc}")
        metadata.append(row.to_dict())
    print(f"  [RAG] Encoding {len(documents)} API entries...", flush=True)
    embeddings = embed_model.encode(documents, convert_to_tensor=True,
                                    show_progress_bar=False)

    # ── verifiers (respecting ablation flags) ─────────────────────────────────
    static_ver: OpenROADStaticVerifier = (
        _PassthroughStaticVerifier()    # type: ignore[assignment]
        if no_causal_verifier           # static disabled → use passthrough
        else OpenROADStaticVerifier(rag_api_path)
    )
    causal_ver: CausalVerifier = (
        _PassthroughCausalVerifier()    # type: ignore[assignment]
        if no_causal_verifier
        else CausalVerifier(metadata=metadata)
    )
    llm_ver: Optional[CausalLLMVerifier] = (
        None
        if no_llm_verifier
        else CausalLLMVerifier(api_key=api_key, model=openai_model, fail_open=True)
    )

    active = []
    if not no_causal_verifier:
        active.append("static")
        active.append("causal")
    else:
        active.append("static(passthrough)")
    if not no_llm_verifier:
        active.append("LLM")
    print(f"[Init] Verifiers active: {', '.join(active)}", flush=True)

    # ── code examples (RAGCodePiece.csv) ──────────────────────────────────────
    code_pieces = []
    if rag_code_path and os.path.isfile(rag_code_path):
        cp_df = pd.read_csv(rag_code_path)
        for _, row in cp_df.iterrows():
            desc = str(row.get("Description:", "")).strip()
            code = str(row.get("Code Piece:", "")).strip()
            if desc and code and desc.lower() != "nan" and code.lower() != "nan":
                code_pieces.append({"description": desc, "code": code})
        print(f"[Init] Loaded {len(code_pieces)} code examples.", flush=True)
    else:
        print("[Init] RAGCodePiecePath not set — code examples disabled.", flush=True)

    # ── OpenROAD PTY ──────────────────────────────────────────────────────────
    master_fd, slave_fd = os.openpty()
    oq                  = queue.Queue()
    threading.Thread(
        target=readOpenROADOutput,
        args=(master_fd, oq, "STDOUT", threading.Event()),
        daemon=True,
    ).start()
    print("[Init] OpenROAD PTY ready.", flush=True)

    # ── load dataset ──────────────────────────────────────────────────────────
    all_prompts, all_gt = load_corpus(test_set_path, 0, prompt_file)

    if cases is not None:
        indexed_cases = [(i, all_prompts[i], all_gt[i])
                         for i in cases if i < len(all_prompts)]
    else:
        sliced_p = all_prompts[start_case:] if start_case else all_prompts
        sliced_g = all_gt[start_case:]      if start_case else all_gt
        if num_cases:
            sliced_p = sliced_p[:num_cases]
            sliced_g = sliced_g[:num_cases]
        offset = start_case or 0
        indexed_cases = [(offset + i, p, g)
                         for i, (p, g) in enumerate(zip(sliced_p, sliced_g))]

    print(f"\nLoaded  : {len(indexed_cases)} prompts from {test_set_path}")
    print(f"Model   : {openai_model}")
    print("=" * 70, flush=True)

    # ── result file setup ─────────────────────────────────────────────────────
    tag_suffix = f"_{run_tag}" if run_tag else f"_{ablation_tag}"
    if result_path.endswith(".csv") or result_path.endswith(".xlsx"):
        _result_file = result_path
    else:
        os.makedirs(result_path, exist_ok=True)
        dataset_tag  = os.path.splitext(os.path.basename(test_set_path))[0]
        _result_file = os.path.join(
            result_path,
            f"{openai_model.replace('.', '-')}__causal_corpus_{dataset_tag}{tag_suffix}.xlsx"
        )

    def _clean(s):
        return ILLEGAL_CHARACTERS_RE.sub("", str(s or ""))

    _HEADERS = [
        "prompt", "ground_truth", "chain", "node_apis",
        "generated_code", "causal_verdict", "llm_verdict",
        "openroad_result", "openroad_output",
        "ctrl_actions", "steps_used", "budget_used",
        "latency_s", "openroad_calls", "ablation_mode",
    ]

    if os.path.exists(_result_file) and _result_file.endswith(".xlsx"):
        wb = load_workbook(_result_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        for col, header in enumerate(_HEADERS, 1):
            ws.cell(row=1, column=col, value=header)

    passed = 0
    total  = 0
    total_latency    = 0.0
    total_ora_calls  = 0

    # ── case loop ─────────────────────────────────────────────────────────────
    for case_i, task, gt_code in indexed_cases:
        row    = case_i + 2
        total += 1
        task   = task.strip()
        print(f"\n{'='*60}", flush=True)
        print(f"[{case_i+1}/{len(indexed_cases)}] {task[:100]}", flush=True)

        case_start   = time.time()
        ora_calls    = 0

        # Fresh OpenROAD process per case
        orig_dir = os.getcwd()
        if run_dir:
            os.chdir(run_dir)
        proc = runOpenROADShell(openroad_path, load_design_time, slave_fd, "")
        if run_dir:
            os.chdir(orig_dir)
        time.sleep(load_design_time)
        clearQueue(oq)

        # ── causal agent loop ──────────────────────────────────────────────
        state = CausalAgentState(task=task, max_budget=budget)
        state = run_case(
            task, state,
            api_key=api_key, model=openai_model,
            embed_model=embed_model, metadata=metadata, embeddings=embeddings,
            static_verifier=static_ver, causal_verifier=causal_ver,
            code_pieces=code_pieces,
            llm_verifier=llm_ver,
        )

        # ── run committed code in OpenROAD ────────────────────────────────
        stdout, has_tb = "", True
        if state.committed_code:
            try:
                stdout, has_tb = _send_code(
                    proc, state.committed_code, oq,
                    max_wait_time, command_flush_time,
                )
                ora_calls += 1
            except Exception as exc:
                stdout = f"[OpenROAD error] {exc}"
                has_tb = True
                ora_calls += 1

        case_latency = time.time() - case_start
        total_latency   += case_latency
        total_ora_calls += ora_calls

        ora_pass = bool(state.committed_code) and not has_tb
        if ora_pass:
            passed += 1

        # terminate OpenROAD
        try:
            proc.terminate()
            proc.wait()
        except Exception:
            pass

        status = "PASS" if ora_pass else "FAIL"
        print(f"  [OpenROAD] {status}  latency={case_latency:.1f}s  "
              f"ora_calls={ora_calls}  budget_used={state.budget_used}", flush=True)
        for line in _clean_output(stdout).splitlines()[:10]:
            print(f"    {line}", flush=True)

        # ── verifier verdict strings ───────────────────────────────────────
        sv = state.static_result
        lv = state.llm_result
        causal_verdict = (
            "PASS" if sv and sv.passed
            else f"FAIL(L{sv.layer_failed}): {'; '.join(sv.issues[:2])}"
            if sv else "N/A"
        )
        llm_verdict = (
            f"PASS(conf={lv.confidence:.2f})" if lv and lv.passed
            else f"FAIL(conf={lv.confidence:.2f}): {'; '.join(lv.issues[:1])}"
            if lv else "not_run"
        )
        ctrl_history = " | ".join(
            f"{o.action}:{o.result[:50]}"
            for o in state.observations
            if not o.is_bootstrap
        )

        # ── write row ──────────────────────────────────────────────────────
        row_vals = [
            _clean(task),
            _clean(gt_code),
            " -> ".join(state.chain),
            _clean(state.api_summary),
            _clean(state.committed_code),
            causal_verdict,
            llm_verdict,
            status,
            _clean(_clean_output(stdout)[:2000]),
            _clean(ctrl_history),
            state.step,
            state.budget_used,
            round(case_latency, 2),
            ora_calls,
            ablation_tag,
        ]
        for col, val in enumerate(row_vals, 1):
            ws.cell(row=row, column=col, value=val)
        wb.save(_result_file)

    # ── summary ───────────────────────────────────────────────────────────────
    avg_latency   = total_latency   / max(total, 1)
    avg_ora_calls = total_ora_calls / max(total, 1)

    print(f"\n{'='*70}", flush=True)
    print(f"[{ablation_tag}] OpenROAD: {passed}/{total} passed "
          f"({passed/max(total,1)*100:.1f}%)", flush=True)
    print(f"  Avg latency      : {avg_latency:.1f}s / case", flush=True)
    print(f"  Avg OpenROAD calls: {avg_ora_calls:.2f} / case", flush=True)
    print(f"Results → {_result_file}", flush=True)


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(
        description="Causal agent runner for EDA-Corpus-v2/TestSet.xlsx"
    )
    p.add_argument("--testSetPath",      required=True,
                   help="Path to EDA-Corpus-v2/TestSet.xlsx")
    p.add_argument("--RAGApiPath",       default="RAGData/RAGAPIs.csv")
    p.add_argument("--RAGCodePiecePath", default="RAGData/RAGCodePiece.csv")
    p.add_argument("--openaiKey",        required=True)
    p.add_argument("--model",            default="gpt-4.1-mini")
    p.add_argument("--budget",           type=int, default=6,
                   help="Causal controller action budget per case")
    p.add_argument("--openroadPath",     default="OpenROAD/build/src/openroad")
    p.add_argument("--runDir",           default="src_1_reflector",
                   help="Working dir for OpenROAD shell")
    p.add_argument("--resultPath",       default="result/causal_corpus_run.xlsx",
                   help="Output .xlsx/.csv path, or a directory")
    p.add_argument("--promptFile",       default="",
                   help="Text file with one prompt per line — run only matching prompts")
    p.add_argument("--numCases",         type=int, default=None)
    p.add_argument("--startCase",        type=int, default=None,
                   help="0-indexed case to start from")
    p.add_argument("--cases",            type=int, nargs="+", default=None,
                   help="Specific 0-indexed case numbers (e.g. --cases 2 6)")
    p.add_argument("--loadDesignTime",   type=int,   default=5)
    p.add_argument("--maxWaitTime",      type=int,   default=120)
    p.add_argument("--commandFlushTime", type=float, default=0.1)
    p.add_argument("--runTag",           default="",
                   help="Optional tag appended to result filename")
    # ── ablation flags ────────────────────────────────────────────────────────
    p.add_argument("--noLLMVerifier",    action="store_true",
                   help="Disable LLM (GPT-4o) semantic verifier — run static+causal only")
    p.add_argument("--noCausalVerifier", action="store_true",
                   help="Disable causal+LLM verifiers — run static verifier only")
    args = p.parse_args()

    def _abs(s):
        return s if os.path.isabs(s) else os.path.join(_ROOT, s)

    Run(
        test_set_path      = _abs(args.testSetPath),
        result_path        = _abs(args.resultPath),
        rag_api_path       = _abs(args.RAGApiPath),
        rag_code_path      = _abs(args.RAGCodePiecePath),
        openroad_path      = _abs(args.openroadPath),
        api_key            = args.openaiKey,
        openai_model       = args.model,
        budget             = args.budget,
        load_design_time   = args.loadDesignTime,
        max_wait_time      = args.maxWaitTime,
        command_flush_time = args.commandFlushTime,
        num_cases          = args.numCases,
        start_case         = args.startCase,
        cases              = args.cases,
        run_dir            = _abs(args.runDir) if args.runDir else "",
        run_tag            = args.runTag,
        prompt_file        = args.promptFile,
        no_llm_verifier    = args.noLLMVerifier,
        no_causal_verifier = args.noCausalVerifier,
    )


if __name__ == "__main__":
    main()
