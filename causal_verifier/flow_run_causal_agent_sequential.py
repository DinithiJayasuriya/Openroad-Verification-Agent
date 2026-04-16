"""flow_run_causal_agent_sequential.py — Flow-task sequential causal agentic eval.

High-level idea (two-level causal structure):
  TOP LEVEL  (blue boxes):  a flow prompt is decomposed into an ordered ActionGraph
                            of flow tools  (e.g. floorplanning → io_placement).
  BOTTOM LEVEL (orange nodes): each action is a full causal chain that is solved by
                            the existing run_case() pipeline from run_causal_agent.py
                            (chain extraction → node RAG → generation → verification).

Pipeline per task
-----------------
  1. TaskDecomposer  (GPT LLM call) : prompt → ActionGraph  [N tool nodes in order]
  2. For each ActionNode in order:
       a. Build a sub-task string from the node description + tool hints
       b. run_case() — full causal bootstrap + controller loop
       c. Execute committed code in the SAME OpenROAD process
       d. If OpenROAD FAILS → skip all remaining actions for this task
       e. Pass accepted code from prior actions as context to the next action

Dataset format (single column):
    prompt   ← one natural-language flow task per row

Usage:
  conda activate prompt
  python causal_verifier_4_2/flow_run_causal_agent_sequential.py \\
    --openaiKey    sk-... \\
    --dataset      path/to/flow_tasks.xlsx \\
    --RAGApiPath   RAGData/RAGAPIs.csv \\
    --RAGCodePiecePath RAGData/RAGCodePiece.csv \\
    --openroadPath OpenROAD/build/src/openroad \\
    --runDir       src_1_reflector \\
    --resultPath   result/ \\
    --budget       6 \\
    --numCases     10
"""

import argparse
import importlib.util as _ilu
import os
import queue
import re
import sys
import threading
import time
from typing import List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from sentence_transformers import SentenceTransformer

# ── path setup ────────────────────────────────────────────────────────────────
_ROOT        = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_CAUSAL_DIR  = os.path.dirname(os.path.abspath(__file__))          # causal_verifier_4_2/
_NEURIPS_DIR = os.path.join(_ROOT, "causal_verifier", "NeurIPS")   # causal_verifier/NeurIPS/

sys.path.insert(0, os.path.join(_ROOT, "src_1_reflector"))
sys.path.insert(0, os.path.join(_ROOT, "src_1_agentic"))
sys.path.insert(0, _CAUSAL_DIR)
sys.path.insert(0, _NEURIPS_DIR)

from util import (
    runOpenROADShell, sendCommandOpenROAD, processCodeString,
    clearQueue, readOpenROADOutput,
)
from verifier            import OpenROADStaticVerifier
from causal_state        import CausalAgentState
from causal_verifier     import CausalVerifier
from llm_verifier        import CausalLLMVerifier
from structured_rag_gate import StructuredRAGGate
from run_causal_agent    import run_case, _clean_output

# ── NeurIPS flow imports ──────────────────────────────────────────────────────
from flow_task_decomposer import TaskDecomposer, ActionGraph, ActionNode
from flow_tool_library    import TOOL_LIBRARY, FlowToolDef


# ─────────────────────────────────────────────────────────────────────────────
# Sub-task prompt builder
# ─────────────────────────────────────────────────────────────────────────────

def _action_to_task(node: ActionNode, tool_def: Optional[FlowToolDef]) -> str:
    """Convert one ActionNode into a task string suitable for run_case().

    Uses the natural-language description from the decomposer as the core,
    and appends method-level hints so the causal chain extractor has precise
    signal about which API sequence to generate.
    """
    desc = node.description.strip()

    if tool_def is None:
        return desc

    # Build a hint string from the tool's required method sequence
    required_methods = [m.name for m in tool_def.default_sequence if m.required]

    # If there are exclusive groups (no default_sequence), pick the first option
    if not required_methods and tool_def.exclusive_groups:
        required_methods = tool_def.exclusive_groups[0].options[0]

    if required_methods:
        method_hint = ", ".join(f"{tool_def.var_name}.{m}()" for m in required_methods)
        return (
            f"{desc}. "
            f"Use design.{tool_def.getter}() to get the {tool_def.tool_type} object, "
            f"then call: {method_hint}."
        )

    # Tool with no defined method sequence — just hint the getter
    return (
        f"{desc}. "
        f"Use design.{tool_def.getter}() to get the {tool_def.tool_type} object."
    )


# ─────────────────────────────────────────────────────────────────────────────
# OpenROAD execution helper
# ─────────────────────────────────────────────────────────────────────────────

def _send_code(proc, code: str, oq: queue.Queue,
               max_wait: int, flush_time: float) -> Tuple[str, bool]:
    cmd = processCodeString(code)
    return sendCommandOpenROAD(proc, cmd, oq, max_wait, flush_time)


# ─────────────────────────────────────────────────────────────────────────────
# Main eval loop
# ─────────────────────────────────────────────────────────────────────────────

def Run(
    dataset_path:        str,
    result_path:         str,
    rag_api_path:        str,
    rag_code_path:       str,
    openroad_path:       str,
    api_key:             str,
    openai_model:        str           = "gpt-4.1-mini",
    budget:              int           = 6,
    load_design_time:    int           = 5,
    max_wait_time:       int           = 120,
    command_flush_time:  float         = 0.1,
    num_cases:           Optional[int] = None,
    start_case:          Optional[int] = None,
    run_dir:             str           = "",
    run_tag:             str           = "",
    rag_structured_path: str           = "",
):
    # ── RAG setup ─────────────────────────────────────────────────────────────
    print("\nLoading embedding model...", flush=True)
    embed_model = SentenceTransformer("all-MiniLM-L6-v2")

    rag_df    = pd.read_csv(rag_api_path)
    metadata  = []
    documents = []
    for _, row in rag_df.iterrows():
        desc = str(row.get("Description:", "")).strip()
        if not desc or desc.lower() == "nan":
            continue
        documents.append(f"OpenROAD Python API Description:{desc}")
        metadata.append(row.to_dict())
    print(f"  [RAG] Encoding {len(documents)} API entries...", flush=True)
    embeddings = embed_model.encode(documents, convert_to_tensor=True,
                                    show_progress_bar=False)

    # ── Structured RAG gate ───────────────────────────────────────────────────
    structured_df = None
    gate          = None
    if rag_structured_path and os.path.isfile(rag_structured_path):
        structured_df = pd.read_csv(rag_structured_path)
        gate          = StructuredRAGGate(rag_structured_path)
        print(f"[Init] StructuredRAGGate loaded "
              f"({len(structured_df)} rows).", flush=True)
    else:
        print("[Init] RAGStructuredPath not set — node gate disabled.", flush=True)

    # ── Verifiers ─────────────────────────────────────────────────────────────
    static_ver = OpenROADStaticVerifier(rag_api_path)
    causal_ver = CausalVerifier(metadata=metadata)
    llm_ver    = CausalLLMVerifier(api_key=api_key, model=openai_model, fail_open=True)
    print("[Init] Static + causal + LLM verifiers ready.", flush=True)

    # ── Code examples ─────────────────────────────────────────────────────────
    code_pieces = []
    if rag_code_path and os.path.isfile(rag_code_path):
        cp_df = pd.read_csv(rag_code_path)
        for _, row in cp_df.iterrows():
            desc = str(row.get("Description:", "")).strip()
            code = str(row.get("Code Piece:", "")).strip()
            if desc and code and desc.lower() != "nan" and code.lower() != "nan":
                code_pieces.append({"description": desc, "code": code})
        print(f"[Init] Loaded {len(code_pieces)} code examples.", flush=True)

    # ── TaskDecomposer ────────────────────────────────────────────────────────
    decomposer = TaskDecomposer(openai_key=api_key, model=openai_model)
    print(f"[Init] TaskDecomposer ready (model={openai_model}).", flush=True)

    # ── OpenROAD PTY ──────────────────────────────────────────────────────────
    master_fd, slave_fd = os.openpty()
    oq                  = queue.Queue()
    threading.Thread(
        target=readOpenROADOutput,
        args=(master_fd, oq, "STDOUT", threading.Event()),
        daemon=True,
    ).start()
    print("[Init] OpenROAD PTY ready.", flush=True)

    # ── Load dataset ──────────────────────────────────────────────────────────
    ext = os.path.splitext(dataset_path)[1].lower()
    df  = pd.read_excel(dataset_path) if ext in (".xlsx", ".xls") else pd.read_csv(dataset_path)
    if "prompt" not in df.columns:
        print(f"ERROR: no 'prompt' column. Available: {list(df.columns)}")
        sys.exit(1)

    prompts = df["prompt"].dropna().tolist()
    if start_case:
        prompts = prompts[start_case:]
    if num_cases:
        prompts = prompts[:num_cases]
    print(f"[Init] {len(prompts)} prompt(s) loaded from {dataset_path}", flush=True)

    # ── Result workbook ───────────────────────────────────────────────────────
    os.makedirs(result_path, exist_ok=True)
    dataset_tag = os.path.splitext(os.path.basename(dataset_path))[0]
    tag_suffix  = f"_{run_tag}" if run_tag else ""
    result_file = os.path.join(
        result_path,
        f"{openai_model.replace('.', '-')}__flow_causal_{dataset_tag}{tag_suffix}.xlsx"
    )

    if os.path.exists(result_file):
        wb = load_workbook(result_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        headers = [
            "prompt",           # A: original flow task
            "action_graph",     # B: decomposed ordering e.g. floorplanning → io_placement
            "num_actions",      # C: number of tool actions
            "action1_task",     # D: sub-task sent to run_case for action 1
            "action1_code",     # E
            "action1_verdict",  # F: causal verifier verdict
            "action1_result",   # G: PASS / FAIL / SKIP (OpenROAD)
            "action1_output",   # H
            "action2_task",     # I
            "action2_code",     # J
            "action2_verdict",  # K
            "action2_result",   # L
            "action2_output",   # M
            "overall",          # N
            "action1_lessons",  # O
            "action2_lessons",  # P
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

    def _clean(s: str) -> str:
        return ILLEGAL_CHARACTERS_RE.sub("", str(s or ""))

    # Column layout helpers (1-indexed, supports up to 2 actions for now)
    _ACTION_COLS = {
        1: {"task": 4,  "code": 5,  "verdict": 6,  "result": 7,  "output": 8,  "lessons": 15},
        2: {"task": 9,  "code": 10, "verdict": 11, "result": 12, "output": 13, "lessons": 16},
    }

    passed_all = 0
    total      = 0

    for case_i, flow_prompt in enumerate(prompts):
        flow_prompt = str(flow_prompt).strip()
        row          = (start_case or 0) + case_i + 2
        total       += 1

        print(f"\n{'='*60}", flush=True)
        print(f"CASE {case_i+1}/{len(prompts)}: {flow_prompt}", flush=True)

        # ── Step 1: Decompose flow prompt → ActionGraph ────────────────────────
        action_graph: Optional[ActionGraph] = decomposer.decompose(flow_prompt)

        if action_graph is None or not action_graph.tool_nodes:
            print("  [decompose] FAIL — could not decompose into flow actions", flush=True)
            ws.cell(row=row, column=1, value=_clean(flow_prompt))
            ws.cell(row=row, column=2, value="DECOMPOSE_FAIL")
            ws.cell(row=row, column=3, value=0)
            ws.cell(row=row, column=14, value="DECOMPOSE_FAIL")
            wb.save(result_file)
            continue

        tool_nodes = action_graph.tool_nodes
        print(f"  [decompose] {action_graph.ordering_summary()} "
              f"({len(tool_nodes)} action(s))", flush=True)

        ws.cell(row=row, column=1, value=_clean(flow_prompt))
        ws.cell(row=row, column=2, value=_clean(action_graph.ordering_summary()))
        ws.cell(row=row, column=3, value=len(tool_nodes))

        # ── Step 2: Fresh OpenROAD process for this task ───────────────────────
        orig_dir = os.getcwd()
        if run_dir:
            os.chdir(run_dir)
        proc = runOpenROADShell(openroad_path, load_design_time, slave_fd, "")
        if run_dir:
            os.chdir(orig_dir)
        time.sleep(load_design_time)
        clearQueue(oq)

        accepted_codes: List[str] = []
        action_results: List[bool] = []

        # ── Step 3: Run each tool action through the causal pipeline ───────────
        for action_idx, node in enumerate(tool_nodes[:2]):   # cap at 2 for now
            action_n  = action_idx + 1
            tool_def  = TOOL_LIBRARY.get(node.tool_id) if node.tool_id else None
            sub_task  = _action_to_task(node, tool_def)

            print(f"\n  ── Action {action_n}/{len(tool_nodes)}: {sub_task[:100]}",
                  flush=True)

            # Skip if previous action failed in OpenROAD
            if action_idx > 0 and not action_results[-1]:
                print(f"  [Action {action_n}] SKIP — action {action_n-1} failed",
                      flush=True)
                if action_n in _ACTION_COLS:
                    c = _ACTION_COLS[action_n]
                    ws.cell(row=row, column=c["task"],    value=_clean(sub_task))
                    ws.cell(row=row, column=c["result"],  value="SKIP")
                    ws.cell(row=row, column=c["verdict"], value="SKIP")
                action_results.append(False)
                continue

            # Inject prior accepted code as session context
            if accepted_codes:
                ctx_block = "\n\n".join(
                    f"# --- Action {i+1} (already executed) ---\n{c}"
                    for i, c in enumerate(accepted_codes)
                )
                task_with_ctx = (
                    f"{sub_task}\n\n"
                    f"[Context] The following actions have already executed in this "
                    f"OpenROAD session.\n"
                    f"```python\n{ctx_block}\n```\n\n"
                    f"IMPORTANT — Variable Scope Rule:\n"
                    f"Only variables assigned at the TOP LEVEL of the above code "
                    f"are guaranteed to be in scope. "
                    f"Re-acquire objects you need (e.g. block = design.getBlock()). "
                    f"Do NOT re-run lines that mutate the design state."
                )
            else:
                task_with_ctx = sub_task

            # ── Causal agent loop ──────────────────────────────────────────────
            state = CausalAgentState(task=task_with_ctx, max_budget=budget)
            state = run_case(
                task_with_ctx, state,
                api_key=api_key, model=openai_model,
                embed_model=embed_model, metadata=metadata, embeddings=embeddings,
                static_verifier=static_ver, causal_verifier=causal_ver,
                code_pieces=code_pieces,
                llm_verifier=llm_ver,
                gate=gate,
                structured_df=structured_df,
            )

            # Causal verdict string
            sv = state.static_result
            if sv:
                causal_verdict = (
                    "PASS" if sv.passed
                    else f"FAIL(L{sv.layer_failed}): {'; '.join(sv.issues[:2])}"
                )
            else:
                causal_verdict = "N/A"

            # ── Execute in OpenROAD ────────────────────────────────────────────
            stdout, has_tb = "", True
            if state.committed_code:
                try:
                    stdout, has_tb = _send_code(
                        proc, state.committed_code, oq,
                        max_wait_time, command_flush_time,
                    )
                except Exception as exc:
                    stdout  = f"[exec error] {exc}"
                    has_tb  = True

            passed = bool(state.committed_code) and not has_tb
            action_results.append(passed)
            if passed:
                accepted_codes.append(state.committed_code)

            ora_label = "PASS" if passed else "FAIL"
            print(f"  [Action {action_n} OpenROAD] {ora_label}", flush=True)
            if not passed and stdout:
                print(f"  [Action {action_n} error]\n{stdout.strip()[:300]}", flush=True)

            # Write action columns
            if action_n in _ACTION_COLS:
                c = _ACTION_COLS[action_n]
                ws.cell(row=row, column=c["task"],    value=_clean(sub_task))
                ws.cell(row=row, column=c["code"],    value=_clean(state.committed_code))
                ws.cell(row=row, column=c["verdict"], value=_clean(causal_verdict))
                ws.cell(row=row, column=c["result"],  value=ora_label)
                ws.cell(row=row, column=c["output"],  value=_clean(stdout[:1500]))
                ws.cell(row=row, column=c["lessons"],
                        value=" | ".join(state.lessons))

        # ── Terminate OpenROAD ─────────────────────────────────────────────────
        try:
            proc.terminate()
            proc.wait()
        except Exception:
            pass

        all_passed = bool(action_results) and all(action_results)
        if all_passed:
            passed_all += 1

        if all_passed:
            overall = "PASS"
        elif not action_results:
            overall = "DECOMPOSE_FAIL"
        else:
            first_fail = next((i for i, r in enumerate(action_results) if not r), None)
            overall = f"ACTION{first_fail+1}_FAIL" if first_fail is not None else "FAIL"

        ws.cell(row=row, column=14, value=overall)
        wb.save(result_file)
        print(f"  [Overall] {overall}", flush=True)

    print(f"\n{'='*60}", flush=True)
    print(f"All actions passed: {passed_all}/{total} = "
          f"{passed_all/max(total,1)*100:.1f}%")
    print(f"Results → {result_file}")


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(
        description="Flow-task sequential causal agentic eval: "
                    "decompose flow prompt → run each action through causal pipeline"
    )
    p.add_argument("--openaiKey",         required=True)
    p.add_argument("--model",             default="gpt-4.1-mini")
    p.add_argument("--dataset",           required=True,
                   help="Path to .xlsx or .csv with a 'prompt' column")
    p.add_argument("--resultPath",        default="result/")
    p.add_argument("--RAGApiPath",        default="RAGData/RAGAPIs.csv")
    p.add_argument("--RAGStructuredPath", default="RAGData/RAGAPIs_structured.csv")
    p.add_argument("--RAGCodePiecePath",  default="RAGData/RAGCodePiece.csv")
    p.add_argument("--openroadPath",      default="OpenROAD/build/src/openroad")
    p.add_argument("--runDir",            default="src_1_reflector",
                   help="Working dir for OpenROAD shell")
    p.add_argument("--budget",            type=int, default=6,
                   help="Causal controller action budget per action")
    p.add_argument("--numCases",          type=int, default=None)
    p.add_argument("--startCase",         type=int, default=None)
    p.add_argument("--loadDesignTime",    type=int, default=5)
    p.add_argument("--maxWaitTime",       type=int, default=120)
    p.add_argument("--commandFlushTime",  type=float, default=0.1)
    p.add_argument("--runTag",            default="")
    args = p.parse_args()

    def _abs(s: str) -> str:
        return s if (not s or os.path.isabs(s)) else os.path.join(_ROOT, s)

    Run(
        dataset_path        = _abs(args.dataset),
        result_path         = _abs(args.resultPath),
        rag_api_path        = _abs(args.RAGApiPath),
        rag_structured_path = _abs(args.RAGStructuredPath),
        rag_code_path       = _abs(args.RAGCodePiecePath),
        openroad_path       = _abs(args.openroadPath),
        api_key             = args.openaiKey,
        openai_model        = args.model,
        budget              = args.budget,
        load_design_time    = args.loadDesignTime,
        max_wait_time       = args.maxWaitTime,
        command_flush_time  = args.commandFlushTime,
        num_cases           = args.numCases,
        start_case          = args.startCase,
        run_dir             = _abs(args.runDir) if args.runDir else "",
        run_tag             = args.runTag,
    )


if __name__ == "__main__":
    main()
