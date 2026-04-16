"""tool_in_loop_baseline.py — Tool-in-loop baseline for multi-step sequential eval.

No causal chain, no RAG, no static/causal/LLM verifiers.
Loop: LLM generates code → run in OpenROAD → if fail, pass error back to LLM → repeat.

Shares the same CLI interface as run_causal_agent_sequential.py so results are
directly comparable (same dataset, --level, --budget, --numCases, etc.).

Usage:
  python causal_verifier_4_2/tool_in_loop_baseline.py \\
    --openaiKey sk-... \\
    --actionDBPath causal_verifier/adversarial_evolved_100.xlsx \\
    --sheet sweet_spot \\
    --openroadPath OpenROAD/build/src/openroad \\
    --runDir src_1_reflector \\
    --resultPath result/ \\
    --level 4 --budget 6 --numCases 50 | tee run_baseline.log
"""

import argparse
import json
import os
import queue
import re
import sys
import threading
import time
import urllib.request
import urllib.error
from typing import List, Optional, Tuple

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

# ── path setup ────────────────────────────────────────────────────────────────
_ROOT      = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_CAUSAL_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(_ROOT, "src_1_reflector"))
sys.path.insert(0, os.path.join(_ROOT, "src_1_agentic"))

from util import (
    runOpenROADShell, sendCommandOpenROAD, processCodeString,
    clearQueue, readOpenROADOutput,
)


# ─────────────────────────────────────────────────────────────────────────────
# LLM helpers
# ─────────────────────────────────────────────────────────────────────────────

_GEN_SYSTEM = """\
You are an OpenROAD Python API expert. Generate executable Python code for \
the OpenROAD interactive shell.

Rules:
- The variables `design` (openroad.Design) and `tech` (odb.dbTech) are pre-available.
- Do NOT import openroad or redefine design/tech.
- To access the design block: block = design.getBlock()
- Output ONLY a Python code block — no explanation, no markdown fences."""


def _call_openai(messages: list, api_key: str, model: str,
                 max_tokens: int = 800) -> str:
    payload = json.dumps({
        "model": model,
        "messages": messages,
        "temperature": 0,
        "max_tokens": max_tokens,
    }).encode()
    req = urllib.request.Request(
        "https://api.openai.com/v1/chat/completions",
        data=payload,
        headers={"Content-Type": "application/json",
                 "Authorization": f"Bearer {api_key}"},
        method="POST",
    )
    for attempt in range(4):
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                body = json.loads(resp.read().decode())
            return body["choices"][0]["message"]["content"].strip()
        except urllib.error.HTTPError as e:
            if e.code == 429:
                wait = 10 * (2 ** attempt)
                print(f"    [rate-limit] waiting {wait}s...", flush=True)
                time.sleep(wait)
            else:
                print(f"    [HTTP {e.code}]", flush=True)
                return ""
        except Exception as exc:
            print(f"    [LLM error] {exc}", flush=True)
            return ""
    return ""


def _extract_code(raw: str) -> str:
    """Strip markdown fences; return plain code."""
    raw = raw.strip()
    if raw.startswith("```"):
        lines = raw.splitlines()
        # drop opening fence (```python or ```) and closing ```
        lines = [ln for ln in lines if not ln.strip().startswith("```")]
        raw = "\n".join(lines).strip()
    return raw


# ─────────────────────────────────────────────────────────────────────────────
# Tool-in-loop core: generate → run → repair × budget
# ─────────────────────────────────────────────────────────────────────────────

def run_tool_in_loop(
    task: str,
    proc,
    oq: queue.Queue,
    api_key: str,
    model: str,
    budget: int,
    max_wait: int,
    flush_time: float,
) -> Tuple[str, bool, str, int]:
    """Run the tool-in-loop repair loop for one step.

    Returns (committed_code, passed, last_output, attempts_used).
    """
    messages = [{"role": "system", "content": _GEN_SYSTEM}]
    messages.append({"role": "user", "content": task})

    committed_code = ""
    passed         = False
    last_output    = ""
    attempts       = 0

    for attempt in range(1, budget + 1):
        attempts = attempt
        # ── generate ──────────────────────────────────────────────────────────
        raw  = _call_openai(messages, api_key, model)
        code = _extract_code(raw)
        if not code:
            print(f"    [attempt {attempt}] LLM returned empty — stopping.", flush=True)
            break

        # ── run in OpenROAD ───────────────────────────────────────────────────
        cmd = processCodeString(code)
        try:
            stdout, has_tb = sendCommandOpenROAD(proc, cmd, oq, max_wait, flush_time)
        except Exception as exc:
            stdout = f"[OpenROAD error] {exc}"
            has_tb = True

        last_output = stdout
        print(f"    [attempt {attempt}] {'PASS' if not has_tb else 'FAIL'}", flush=True)

        if not has_tb:
            # success
            committed_code = code
            passed         = True
            break

        # ── repair: append error to conversation and loop ─────────────────────
        # Append what the assistant generated
        messages.append({"role": "assistant", "content": f"```python\n{code}\n```"})
        # Append the OpenROAD error as user feedback
        error_snippet = stdout.strip()[-1500:] if stdout else "OpenROAD returned an error."
        repair_msg = (
            f"The code failed in OpenROAD with this error:\n\n"
            f"```\n{error_snippet}\n```\n\n"
            f"Fix the code to resolve this error and correctly complete the task. "
            f"Output ONLY the corrected Python code block."
        )
        messages.append({"role": "user", "content": repair_msg})

        if attempt == budget:
            # budget exhausted — keep last code as best effort
            committed_code = code

    return committed_code, passed, last_output, attempts


# ─────────────────────────────────────────────────────────────────────────────
# Case loader (identical to run_causal_agent_sequential.py)
# ─────────────────────────────────────────────────────────────────────────────

def load_cases(path: str, level: int,
               sheet: Optional[str] = None) -> List[Tuple[str, List[str]]]:
    wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws   = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    cases = []
    for row in rows[1:]:
        if not row[0] or not str(row[0]).strip():
            continue
        complex_prompt = str(row[0]).strip()
        steps = []
        for col_idx in range(1, level + 1):
            val = row[col_idx] if len(row) > col_idx else None
            steps.append(str(val).strip() if val and str(val).strip() else "")
        cases.append((complex_prompt, steps))
    return cases


# ─────────────────────────────────────────────────────────────────────────────
# Main eval loop
# ─────────────────────────────────────────────────────────────────────────────

def Run(
    action_db_path:     str,
    result_path:        str,
    openroad_path:      str,
    api_key:            str,
    openai_model:       str           = "gpt-4.1-mini",
    level:              int           = 2,
    budget:             int           = 6,
    load_design_time:   int           = 5,
    max_wait_time:      int           = 120,
    command_flush_time: float         = 0.1,
    num_cases:          Optional[int] = None,
    start_case:         Optional[int] = None,
    cases:              Optional[List[int]] = None,
    run_dir:            str           = "",
    run_tag:            str           = "",
    sheet:              Optional[str] = None,
):
    # ── OpenROAD PTY ──────────────────────────────────────────────────────────
    master_fd, slave_fd = os.openpty()
    oq                  = queue.Queue()
    threading.Thread(
        target=readOpenROADOutput,
        args=(master_fd, oq, "STDOUT", threading.Event()),
        daemon=True,
    ).start()
    print("[Init] OpenROAD PTY ready.", flush=True)

    # ── load + slice cases ────────────────────────────────────────────────────
    all_cases = load_cases(action_db_path, level, sheet=sheet)
    if cases is not None:
        indexed_cases = [(i, all_cases[i][0], all_cases[i][1])
                         for i in cases if i < len(all_cases)]
    else:
        if start_case is not None:
            all_cases = all_cases[start_case:]
        if num_cases is not None:
            all_cases = all_cases[:num_cases]
        offset = start_case or 0
        indexed_cases = [(offset + i, cp, steps)
                         for i, (cp, steps) in enumerate(all_cases)]
    print(f"[Init] Level={level}, budget={budget}, "
          f"running {len(indexed_cases)} case(s).", flush=True)

    # ── result workbook ───────────────────────────────────────────────────────
    os.makedirs(result_path, exist_ok=True)
    dataset_tag  = os.path.splitext(os.path.basename(action_db_path))[0]
    tag_suffix   = f"_{run_tag}" if run_tag else ""
    _result_file = os.path.join(
        result_path,
        f"{openai_model.replace('.', '-')}__tool_loop_L{level}_{dataset_tag}{tag_suffix}.xlsx"
    )

    if os.path.exists(_result_file):
        wb = load_workbook(_result_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        _FIELDS = ["prompt", "code", "attempts", "result", "output"]
        headers = {1: "complex_prompt"}
        col = 2
        for n in range(1, level + 1):
            for field in _FIELDS:
                headers[col] = f"step{n}_{field}"
                col += 1
        headers[col] = "overall"; col += 1
        for c, h in headers.items():
            ws.cell(row=1, column=c, value=h)

    _FIELDS = ["prompt", "code", "attempts", "result", "output"]

    def _col(step_n: int, field: str) -> int:
        return (step_n - 1) * len(_FIELDS) + _FIELDS.index(field) + 2

    def _overall_col() -> int:
        return level * len(_FIELDS) + 2

    def _clean(s):
        return ILLEGAL_CHARACTERS_RE.sub("", str(s or ""))

    passed_all = 0
    total      = 0
    total_openroad_calls = 0

    # ── case loop ─────────────────────────────────────────────────────────────
    for case_i, complex_prompt, steps in indexed_cases:
        row   = case_i + 2
        total += 1
        print(f"\n{'='*60}", flush=True)
        print(f"CASE {case_i+1}/{len(indexed_cases)}  (level={level}, {len(steps)} steps)",
              flush=True)

        # Fresh OpenROAD process for each case
        orig_dir = os.getcwd()
        if run_dir:
            os.chdir(run_dir)
        proc = runOpenROADShell(openroad_path, load_design_time, slave_fd, "")
        if run_dir:
            os.chdir(orig_dir)
        time.sleep(load_design_time)
        clearQueue(oq)

        accepted_codes: List[str] = []
        step_results:   List[bool] = []
        step_outputs:   List[str]  = []
        step_codes:     List[str]  = []
        step_attempts:  List[int]  = []

        for step_idx, raw_prompt in enumerate(steps):
            step_n = step_idx + 1
            print(f"\n  ── Step {step_n}/{len(steps)}: {raw_prompt[:100]}", flush=True)

            # Skip if any prior step failed
            if step_idx > 0 and not step_results[-1]:
                print(f"  [Step {step_n}] SKIP — step {step_n-1} failed", flush=True)
                step_results.append(False)
                step_outputs.append("")
                step_codes.append("")
                step_attempts.append(0)
                continue

            # Build task prompt: include accepted prior-step code as context
            if accepted_codes:
                ctx_block = "\n\n".join(
                    f"# --- Step {i+1} (already executed) ---\n{c}"
                    for i, c in enumerate(accepted_codes)
                )
                task = (
                    f"{raw_prompt}\n\n"
                    f"[Context] The following steps have already executed in this "
                    f"OpenROAD session:\n"
                    f"```python\n{ctx_block}\n```\n\n"
                    f"Only variables assigned at the TOP LEVEL of the above code are "
                    f"guaranteed to be in scope. Always re-acquire objects you need "
                    f"(e.g. block = design.getBlock()) at the top of your new code. "
                    f"Do NOT re-run lines that mutate the design state."
                )
            else:
                task = raw_prompt

            committed_code, passed, stdout, attempts = run_tool_in_loop(
                task=task,
                proc=proc,
                oq=oq,
                api_key=api_key,
                model=openai_model,
                budget=budget,
                max_wait=max_wait_time,
                flush_time=command_flush_time,
            )
            total_openroad_calls += attempts

            step_results.append(passed)
            step_outputs.append(stdout)
            step_codes.append(committed_code)
            step_attempts.append(attempts)

            if passed:
                accepted_codes.append(committed_code)
            print(f"  [Step {step_n}] {'PASS' if passed else 'FAIL'} "
                  f"(attempts={attempts})", flush=True)

        # terminate OpenROAD
        try:
            proc.terminate()
            proc.wait()
        except Exception:
            pass

        all_passed = all(step_results)
        if all_passed:
            passed_all += 1

        overall = "PASS" if all_passed else (
            f"STEP{next(i for i, r in enumerate(step_results) if not r)+1}_FAIL"
        )
        print(f"  [Overall] {overall}", flush=True)

        # ── write Excel ────────────────────────────────────────────────────────
        ws.cell(row=row, column=1, value=_clean(complex_prompt))
        for step_idx, (raw_prompt, code, attempts, passed, stdout) in enumerate(
            zip(steps, step_codes, step_attempts, step_results, step_outputs)
        ):
            step_n = step_idx + 1
            ws.cell(row=row, column=_col(step_n, "prompt"),   value=raw_prompt)
            ws.cell(row=row, column=_col(step_n, "code"),     value=_clean(code))
            ws.cell(row=row, column=_col(step_n, "attempts"), value=attempts)
            ws.cell(row=row, column=_col(step_n, "result"),
                    value="PASS" if passed else ("SKIP" if attempts == 0 else "FAIL"))
            ws.cell(row=row, column=_col(step_n, "output"),   value=_clean(stdout[:2000]))

        ws.cell(row=row, column=_overall_col(), value=overall)
        wb.save(_result_file)

    print(f"\n{'='*60}", flush=True)
    print(f"All steps passed: {passed_all}/{total} = "
          f"{passed_all/max(total,1)*100:.1f}%", flush=True)
    print(f"Total OpenROAD calls: {total_openroad_calls}  "
          f"(avg {total_openroad_calls/max(total,1):.2f}/case)", flush=True)
    print(f"Results → {_result_file}", flush=True)


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(
        description="Tool-in-loop baseline: LLM generates, OpenROAD runs, error fed back."
    )
    p.add_argument("--openaiKey",        required=True)
    p.add_argument("--model",            default="gpt-4.1-mini")
    p.add_argument("--actionDBPath",     default="src_1_reflector/action_db.xlsx")
    p.add_argument("--resultPath",       default="result/")
    p.add_argument("--RAGApiPath",       default="",
                   help="Unused — kept for CLI compatibility with causal pipeline")
    p.add_argument("--openroadPath",     default="OpenROAD/build/src/openroad")
    p.add_argument("--runDir",           default="src_1_reflector")
    p.add_argument("--level",            type=int, default=2)
    p.add_argument("--budget",           type=int, default=6,
                   help="Max OpenROAD execution attempts per step")
    p.add_argument("--loadDesignTime",   type=int,   default=5)
    p.add_argument("--maxWaitTime",      type=int,   default=120)
    p.add_argument("--commandFlushTime", type=float, default=0.1)
    p.add_argument("--numCases",         type=int,   default=None)
    p.add_argument("--startCase",        type=int,   default=None)
    p.add_argument("--cases",            type=int, nargs="+", default=None)
    p.add_argument("--runTag",           default="")
    p.add_argument("--sheet",            default=None)
    args = p.parse_args()

    def _abs(s):
        return s if (not s or os.path.isabs(s)) else os.path.join(_ROOT, s)

    Run(
        action_db_path      = _abs(args.actionDBPath),
        result_path         = _abs(args.resultPath),
        openroad_path       = _abs(args.openroadPath),
        api_key             = args.openaiKey,
        openai_model        = args.model,
        level               = args.level,
        budget              = args.budget,
        load_design_time    = args.loadDesignTime,
        max_wait_time       = args.maxWaitTime,
        command_flush_time  = args.commandFlushTime,
        num_cases           = args.numCases,
        start_case          = args.startCase,
        cases               = args.cases,
        run_dir             = _abs(args.runDir) if args.runDir else "",
        run_tag             = args.runTag,
        sheet               = args.sheet,
    )


if __name__ == "__main__":
    main()
