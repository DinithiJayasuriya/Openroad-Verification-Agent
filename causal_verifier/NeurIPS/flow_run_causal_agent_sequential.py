"""flow_run_causal_agent_sequential.py — Sequential flow-task agentic eval.

Same structure as run_causal_agent_sequential.py but wired to the flow pipeline:

  Bootstrap (free, no budget cost):
    1. flow_decompose  — task → ActionGraph       (TaskDecomposer LLM call)
    2. flow_extract    — ActionGraph → MultiActionChains (FlowMultiChainExtractor)
    3. flow_generate   — MultiActionChains → code  (LLM with constraint prompt)
    4. flow_l4a_verify — code → L4aSnapshot       (FlowL4aVerifier AST check)

  Controller loop (budgeted):
    FlowCausalController.decide(state) → FlowControllerDecision
    FlowCausalDispatcher.dispatch(state, decision) → execute action
    repeat until commit_best / stop_fail / budget=0

  OpenROAD execution (once per step, after commit):
    Runs committed_code in the SAME OpenROAD process across steps.

Dataset format (single-prompt column):
    prompt

Usage:
  conda activate prompt
  python causal_verifier/flow_run_causal_agent_sequential.py \\
    --openaiKey    sk-... \\
    --dataset      path/to/flow_tasks.xlsx \\
    --RAGApiPath   RAGData/RAGAPIs.csv \\
    --openroadPath OpenROAD/build/src/openroad \\
    --runDir       src_1_reflector \\
    --resultPath   result/ \\
    --budget       4 \\
    --numCases     10
"""

import argparse
import os
import queue
import re
import sys
import threading
import time
from typing import List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

# ── path setup ────────────────────────────────────────────────────────────────
_ROOT       = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_CAUSAL_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(_ROOT, "src_1_reflector"))
sys.path.insert(0, os.path.join(_ROOT, "src_1_agentic"))
sys.path.insert(0, _CAUSAL_DIR)

# Bootstrap chain_extractor pyc so the package __init__.py can import it
# (silently skipped if the .pyc was compiled for a different Python version)
import importlib.util as _ilu, glob as _glob
_pyc_candidates = _glob.glob(
    os.path.join(_CAUSAL_DIR, "__pycache__", "chain_extractor.cpython-*.pyc")
)
for _pyc in _pyc_candidates:
    try:
        _spec = _ilu.spec_from_file_location("causal_verifier.chain_extractor", _pyc)
        _mod  = _ilu.module_from_spec(_spec)
        sys.modules.setdefault("causal_verifier.chain_extractor", _mod)
        _spec.loader.exec_module(_mod)
        break
    except (ImportError, Exception):
        sys.modules.pop("causal_verifier.chain_extractor", None)
        continue

from util import (
    runOpenROADShell, sendCommandOpenROAD, processCodeString,
    clearQueue, readOpenROADOutput,
)

# ── Pre-load node_retriever with correct package context ──────────────────────
# node_retriever.py uses `from .chain_extractor import CausalChain` (a relative
# import).  When imported as a plain module it fails.  We load both files via
# importlib with __package__ = "causal_verifier" so the relative import resolves
# correctly, then register them under their plain names so that subsequent
# `from node_retriever import …` statements hit the cached module.
#
# This mirrors the golden approach in run_causal_agent_sequential.py: avoid
# importing node_retriever as a standalone module — always go through the
# causal_verifier package namespace.
import importlib.util as _ilu_pre

def _preload_causal_modules():
    if "node_retriever" in sys.modules:
        return
    _cv = _CAUSAL_DIR
    # Step 1: chain_extractor (needed by node_retriever's relative import)
    if "causal_verifier.chain_extractor" not in sys.modules:
        _sp = _ilu_pre.spec_from_file_location(
            "causal_verifier.chain_extractor",
            os.path.join(_cv, "chain_extractor.py"))
        _m = _ilu_pre.module_from_spec(_sp)
        _m.__package__ = "causal_verifier"
        sys.modules["causal_verifier.chain_extractor"] = _m
        sys.modules.setdefault("chain_extractor", _m)
        _sp.loader.exec_module(_m)
    # Step 2: node_retriever
    _sp = _ilu_pre.spec_from_file_location(
        "causal_verifier.node_retriever",
        os.path.join(_cv, "node_retriever.py"))
    _m = _ilu_pre.module_from_spec(_sp)
    _m.__package__ = "causal_verifier"
    sys.modules["causal_verifier.node_retriever"] = _m
    sys.modules["node_retriever"] = _m
    _sp.loader.exec_module(_m)  # from .chain_extractor import CausalChain → resolves

_preload_causal_modules()

import causal_verifier as _cv_pkg   # noqa: F401  (triggers __init__.py — chain_extractor already cached)

from flow_causal_state      import FlowCausalAgentState
from flow_causal_controller import FlowCausalController
from flow_causal_dispatcher import FlowCausalDispatcher


# ── helpers ───────────────────────────────────────────────────────────────────

def _clean(text: str) -> str:
    """Strip ANSI / control chars and illegal xlsx chars."""
    text = re.sub(r'\x1b\][^\x07]*\x07', '', str(text or ""))
    text = re.sub(r'\x1b\[[0-9;]*[a-zA-Z]', '', text)
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', text)
    return ILLEGAL_CHARACTERS_RE.sub("", text.strip())


def _send_code(proc, code: str, oq: queue.Queue,
               max_wait: int, flush_time: float) -> Tuple[str, bool]:
    cmd = processCodeString(code)
    return sendCommandOpenROAD(proc, cmd, oq, max_wait, flush_time)


# ── Flow agent loop (single task) ────────────────────────────────────────────

def flow_run_case(
    task:       str,
    dispatcher: FlowCausalDispatcher,
    controller: FlowCausalController,
    budget:     int,
) -> FlowCausalAgentState:
    """Run the full flow agent loop for one task.

    Bootstrap (free) → controller loop (budgeted) → return final state.
    """
    state = FlowCausalAgentState(task=task, max_budget=budget)

    # ── Bootstrap ──────────────────────────────────────────────────────────────
    dispatcher.bootstrap_flow_decompose(state)
    if state.action_graph is None:
        # Not a flow task — mark as failed immediately
        state.committed      = False
        state.committed_code = ""
        return state

    dispatcher.bootstrap_flow_extract(state)
    dispatcher.bootstrap_flow_generate(state)
    dispatcher.bootstrap_static_verify(state)      # L1-L3 static check
    dispatcher.bootstrap_flow_l4a_verify(state)    # A1-A4 ordering check

    # Early exit if bootstrap already passed both verifiers
    sv_ok  = state.static_result is None or state.static_result.passed
    l4a_ok = state.l4a_result is None or state.l4a_result.passed
    if sv_ok and l4a_ok and state.current_code:
        state.committed      = True
        state.committed_code = state.current_code
        print("  [flow_run_case] PASS at bootstrap", flush=True)
        return state

    # ── Controller loop ─────────────────────────────────────────────────────
    while state.budget_remaining > 0:
        print(
            f"  [flow_run_case] budget={state.budget_remaining}/{budget}  "
            f"checks={state.l4a_result.issue_checks if state.l4a_result else '?'}",
            flush=True,
        )

        decision = controller.decide(state)
        print(
            f"  [flow_run_case] controller → {decision.next_action} "
            f"{'(fallback)' if decision.from_fallback else ''}",
            flush=True,
        )

        dispatcher.dispatch(state, decision)

        if decision.next_action in ("commit_best", "stop_fail"):
            break
        if state.l4a_result and state.l4a_result.passed:
            # Verifier passed after re_generate — commit immediately
            state.committed      = True
            state.committed_code = state.current_code
            print("  [flow_run_case] PASS after re_generate", flush=True)
            break

    # Ensure best code is committed if loop ended without explicit commit
    if not state.committed and state.best_code:
        state.committed      = True
        state.committed_code = state.best_code

    return state


# ── Main eval loop ────────────────────────────────────────────────────────────

def Run(
    dataset_path:       str,
    result_path:        str,
    rag_api_path:       str,
    openroad_path:      str,
    api_key:            str,
    openai_model:       str           = "gpt-4.1-mini",
    budget:             int           = 4,
    load_design_time:   int           = 5,
    max_wait_time:      int           = 120,
    command_flush_time: float         = 0.1,
    num_cases:          Optional[int] = None,
    start_case:         Optional[int] = None,
    run_dir:            str           = "",
    run_tag:            str           = "",
):
    # ── load dataset ──────────────────────────────────────────────────────────
    ext = os.path.splitext(dataset_path)[1].lower()
    df  = pd.read_excel(dataset_path) if ext in (".xlsx", ".xls") else pd.read_csv(dataset_path)
    if "prompt" not in df.columns:
        print(f"ERROR: no 'prompt' column. Available: {list(df.columns)}")
        sys.exit(1)

    prompts = df["prompt"].dropna().tolist()
    if start_case:
        prompts = prompts[start_case:]
    if num_cases:
        prompts = prompts[:num_cases]
    print(f"[Init] {len(prompts)} prompt(s) loaded from {dataset_path}", flush=True)

    # ── flow components ────────────────────────────────────────────────────────
    dispatcher = FlowCausalDispatcher(
        api_key=api_key, rag_api_path=rag_api_path, model=openai_model,
    )
    controller = FlowCausalController(api_key=api_key, model=openai_model)
    print(f"[Init] FlowCausalDispatcher + FlowCausalController ready (model={openai_model})", flush=True)

    # ── OpenROAD PTY ──────────────────────────────────────────────────────────
    master_fd, slave_fd = os.openpty()
    oq                  = queue.Queue()
    stop_event          = threading.Event()
    threading.Thread(
        target=readOpenROADOutput,
        args=(master_fd, oq, "STDOUT", stop_event),
        daemon=True,
    ).start()
    print("[Init] OpenROAD PTY ready.", flush=True)

    # ── result workbook ────────────────────────────────────────────────────────
    os.makedirs(result_path, exist_ok=True)
    dataset_tag  = os.path.splitext(os.path.basename(dataset_path))[0]
    tag_suffix   = f"_{run_tag}" if run_tag else ""
    result_file  = os.path.join(
        result_path,
        f"{openai_model.replace('.', '-')}__flow_seq_{dataset_tag}{tag_suffix}.xlsx"
    )

    if os.path.exists(result_file):
        wb = load_workbook(result_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        headers = [
            "prompt",
            "action_graph",      # ordering summary e.g. global_placement → unplaced_instance_count
            "sandwich",          # True / False
            "l4a_verdict",       # PASS / FAIL: [A1, A2] ...
            "l4a_issues",        # issue descriptions
            "attempts",          # number of generate-verify cycles
            "final_code",        # committed code
            "openroad_result",   # PASS / FAIL / SKIPPED
            "openroad_output",   # first 1000 chars of stdout
            "lessons",           # accumulated lessons
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

    # ── case loop ─────────────────────────────────────────────────────────────
    passed_l4a = 0
    passed_ora = 0
    total      = 0

    for case_i, task in enumerate(prompts):
        task   = str(task).strip()
        row    = (start_case or 0) + case_i + 2
        total += 1

        print(f"\n{'='*60}", flush=True)
        print(f"[{case_i+1}/{len(prompts)}] {task}", flush=True)

        # ── flow agent loop ────────────────────────────────────────────────────
        state = flow_run_case(task, dispatcher, controller, budget)

        # ── collect result fields ─────────────────────────────────────────────
        if state.action_graph is None:
            l4a_verdict  = "SKIP (not a flow task)"
            l4a_issues   = ""
            action_graph = "NOT_FLOW"
            sandwich     = ""
            attempts_n   = 0
        else:
            sv_ok  = state.static_result is None or state.static_result.passed
            l4a_ok = state.l4a_result is not None and state.l4a_result.passed
            action_graph = state.action_graph.ordering_summary()
            sandwich     = str(state.action_graph.sandwich)
            attempts_n   = len([o for o in state.observations if not o.is_bootstrap])

            if sv_ok and l4a_ok:
                l4a_verdict = "PASS"
                l4a_issues  = ""
                passed_l4a += 1
            else:
                parts = []
                if not sv_ok and state.static_result:
                    parts.append(f"L{state.static_result.layer_failed}: "
                                 + "; ".join(state.static_result.issues[:2]))
                if not l4a_ok and state.l4a_result:
                    snap   = state.l4a_result
                    checks = snap.issue_checks
                    parts.append("A: " + ", ".join(checks) if checks else "A: FAIL")
                l4a_verdict = "FAIL: " + " | ".join(parts) if parts else "FAIL"
                l4a_issues  = ""
                if state.l4a_result:
                    l4a_issues = "\n".join(state.l4a_result.issues)

        print(f"  [verdict] {l4a_verdict}  (attempts={attempts_n})", flush=True)

        # ── OpenROAD execution ─────────────────────────────────────────────────
        ora_result = "SKIPPED"
        ora_output = ""

        if openroad_path and state.committed_code:
            print("  [OpenROAD] Running...", flush=True)
            orig_dir = os.getcwd()
            if run_dir:
                os.chdir(run_dir)
            proc = runOpenROADShell(openroad_path, load_design_time, slave_fd, "")
            if run_dir:
                os.chdir(orig_dir)
            time.sleep(load_design_time)
            clearQueue(oq)

            try:
                stdout, has_tb = _send_code(
                    proc, state.committed_code, oq,
                    max_wait_time, command_flush_time,
                )
                ora_result = "FAIL" if has_tb else "PASS"
                ora_output = stdout[:1000]
                if not has_tb:
                    passed_ora += 1
            except Exception as exc:
                ora_result = "FAIL"
                ora_output = f"[exec error] {exc}"

            try:
                proc.terminate()
                proc.wait()
            except Exception:
                pass
            clearQueue(oq)

            print(f"  [OpenROAD] {ora_result}", flush=True)

        # ── write row ──────────────────────────────────────────────────────────
        lessons_str = " | ".join(state.lessons)
        ws.cell(row=row, column=1,  value=_clean(task))
        ws.cell(row=row, column=2,  value=_clean(action_graph))
        ws.cell(row=row, column=3,  value=_clean(sandwich))
        ws.cell(row=row, column=4,  value=_clean(l4a_verdict))
        ws.cell(row=row, column=5,  value=_clean(l4a_issues))
        ws.cell(row=row, column=6,  value=attempts_n)
        ws.cell(row=row, column=7,  value=_clean(state.committed_code))
        ws.cell(row=row, column=8,  value=ora_result)
        ws.cell(row=row, column=9,  value=_clean(ora_output))
        ws.cell(row=row, column=10, value=_clean(lessons_str))
        wb.save(result_file)

    # ── summary ────────────────────────────────────────────────────────────────
    stop_event.set()
    print(f"\n{'='*60}", flush=True)
    print(f"L4a passed   : {passed_l4a}/{total} = {passed_l4a/max(total,1)*100:.1f}%")
    if openroad_path:
        print(f"OpenROAD     : {passed_ora}/{total} = {passed_ora/max(total,1)*100:.1f}%")
    print(f"Results      → {result_file}")


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(
        description="Flow-task sequential agentic eval: decompose → extract → generate → L4a verify"
    )
    p.add_argument("--dataset",          required=True,
                   help="Path to .xlsx or .csv with a 'prompt' column")
    p.add_argument("--openaiKey",        required=True)
    p.add_argument("--model",            default="gpt-4.1-mini")
    p.add_argument("--resultPath",       default="result/")
    p.add_argument("--RAGApiPath",       default="RAGData/RAGAPIs.csv")
    p.add_argument("--openroadPath",     default="OpenROAD/build/src/openroad")
    p.add_argument("--runDir",           default="src_1_reflector",
                   help="Working dir for OpenROAD shell")
    p.add_argument("--budget",           type=int, default=4,
                   help="Controller action budget per task")
    p.add_argument("--numCases",         type=int, default=None)
    p.add_argument("--startCase",        type=int, default=None)
    p.add_argument("--loadDesignTime",   type=int, default=5)
    p.add_argument("--maxWaitTime",      type=int, default=120)
    p.add_argument("--commandFlushTime", type=float, default=0.1)
    p.add_argument("--runTag",           default="",
                   help="Optional tag appended to result filename")
    args = p.parse_args()

    def _abs(s: str) -> str:
        return s if (not s or os.path.isabs(s)) else os.path.join(_ROOT, s)

    Run(
        dataset_path       = _abs(args.dataset),
        result_path        = _abs(args.resultPath),
        rag_api_path       = _abs(args.RAGApiPath),
        openroad_path      = _abs(args.openroadPath),
        api_key            = args.openaiKey,
        openai_model       = args.model,
        budget             = args.budget,
        load_design_time   = args.loadDesignTime,
        max_wait_time      = args.maxWaitTime,
        command_flush_time = args.commandFlushTime,
        num_cases          = args.numCases,
        start_case         = args.startCase,
        run_dir            = _abs(args.runDir) if args.runDir else "",
        run_tag            = args.runTag,
    )


if __name__ == "__main__":
    main()
