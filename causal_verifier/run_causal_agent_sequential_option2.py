"""run_causal_agent_sequential_option2.py — Option 2: chain-extension sequential eval.

Option 2 behaviour
------------------
  Step 1 : Extract causal chain fresh from the sub-prompt, do RAG, generate + verify,
           run in OpenROAD.
  Step N>1: If step N-1 passed in OpenROAD, EXTEND the existing causal chain with the
            new nodes required by sub-prompt N (instead of extracting from scratch).
            Only the NEW edges get RAG retrieval; previously retrieved edges are reused.
            Code is generated with the full extended chain as context.
  A step is skipped entirely if any prior step failed.

Controlled by --level:
  level 1 — runs only step 1 (level_one column)
  level 2 — runs step 1 → step 2 (level_one + level_two columns)
  level 3 — runs all 3 sub-prompts (level_one + level_two + level_three columns)

Each step runs in the SAME OpenROAD session (state carries over).

Usage:
  python causal_verifier/run_causal_agent_sequential_option2.py \\
    --openaiKey sk-... \\
    --actionDBPath src_1_reflector/action_db.xlsx \\
    --RAGApiPath RAGData/RAGAPIs.csv \\
    --RAGCodePiecePath RAGData/RAGCodePiece.csv \\
    --openroadPath OpenROAD/build/src/openroad \\
    --runDir src_1_reflector \\
    --resultPath result/ \\
    --level 3 --budget 6 --numCases 5
"""

import argparse
import os
import queue
import re
import sys
import threading
import time
from typing import List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from sentence_transformers import SentenceTransformer

# ── path setup ────────────────────────────────────────────────────────────────
_ROOT       = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_CAUSAL_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(_ROOT, "src_1_reflector"))
sys.path.insert(0, os.path.join(_ROOT, "src_1_agentic"))
sys.path.insert(0, _CAUSAL_DIR)

from util import (
    runOpenROADShell, sendCommandOpenROAD, processCodeString,
    clearQueue, readOpenROADOutput,
)
from verifier import OpenROADStaticVerifier

from causal_state    import CausalAgentState
from causal_verifier import CausalVerifier
from llm_verifier    import CausalLLMVerifier

# Re-use helpers from run_causal_agent
from run_causal_agent import (
    _clean_output,
    _call_openai, _parse_multi_path_chain, _rag_query_for_edge,
    _rewrite_traversal_chain, bootstrap_causal_rag,
    bootstrap_generate, bootstrap_static_verify,
    run_bootstrap,
    CHAIN_SYSTEM_PROMPT,
)
from controller  import CausalController
from dispatcher  import CausalDispatcher


# ─────────────────────────────────────────────────────────────────────────────
# Option 2 — chain extension helpers
# ─────────────────────────────────────────────────────────────────────────────

_EXTEND_SYSTEM_PROMPT = """\
You are an OpenROAD Python API expert.

You have an existing causal acquisition chain from a prior step.
A new sub-task requires additional objects. Your job is to EXTEND the chain by
adding only the new nodes needed for the new sub-task.

RULES:
- The existing chain nodes are already acquired in the live OpenROAD session — do NOT repeat them.
- Only add NEW terminal nodes not already present in the existing chain.
- Each new path starts from the closest already-acquired node (do not restart from openroad.Design if a shorter path exists through an already-acquired node).
- Use exact OpenROAD Python type names.
- If the new sub-task needs no additional objects beyond what the existing chain provides, output: NO_EXTENSION

OUTPUT FORMAT (no markdown, no JSON):
Existing chain: <shown to you below — do not repeat>
New Path 1 (<NewTerminalType>): <AlreadyAcquiredNode> -> ... -> <NewTerminalType>
New Path 2 (<NewTerminalType>): <AlreadyAcquiredNode> -> ... -> <NewTerminalType>
Action Node: <action>(<TerminalType1>, ...)
"""


def bootstrap_causal_extend(
    state: CausalAgentState,
    prev_state: CausalAgentState,
    api_key: str,
    model: str,
    embed_model,
    metadata: list,
    embeddings,
) -> None:
    """Option 2 bootstrap step 1: extend the previous step's chain for a new sub-task.

    Reuses prev_state.chain and prev_state.edge_apis for already-acquired nodes.
    Only runs RAG for genuinely new edges.  Updates state.chain / state.all_edges /
    state.edge_apis in place.
    """
    prev_chain_str = " -> ".join(prev_state.chain) if prev_state.chain else "(empty)"

    user_msg = (
        f"Existing chain from previous step: {prev_chain_str}\n\n"
        f"New sub-task: {state.task}"
    )

    text = _call_openai(
        messages=[
            {"role": "system", "content": _EXTEND_SYSTEM_PROMPT},
            {"role": "user",   "content": user_msg},
        ],
        api_key=api_key, model=model, max_tokens=300,
    )

    print(f"  [extend/raw] {text[:200]}", flush=True)

    # If the LLM says no new nodes are needed, inherit the previous chain as-is
    if not text.strip() or "NO_EXTENSION" in text.upper():
        print("  [extend] NO_EXTENSION — inheriting previous chain unchanged", flush=True)
        state.chain     = list(prev_state.chain)
        state.paths     = list(prev_state.paths)
        state.all_edges = list(prev_state.all_edges)
        state.edge_apis = list(prev_state.edge_apis)
        state.api_summary = prev_state.api_summary
        state.add_bootstrap_obs("causal_extend", "NO_EXTENSION — reused previous chain")
        return

    # Parse new paths from the extension output
    # Re-use the existing parser but prefix lines to match its "Path N (...): ..." pattern
    normalised = re.sub(r'(?i)^New\s+(Path)', r'\1', text, flags=re.MULTILINE)
    new_paths, action_node = _parse_multi_path_chain(normalised)

    if not new_paths:
        # Fallback: re-extract fresh (LLM output unparseable)
        print("  [extend] parse failed — falling back to fresh extraction", flush=True)
        from run_causal_agent import bootstrap_causal_extract
        bootstrap_causal_extract(state, api_key, model)
        bootstrap_causal_rag(state, embed_model, metadata, embeddings)
        return

    # --- Merge: existing chain nodes + new extension nodes ---
    existing_nodes: List[str] = list(prev_state.chain)
    existing_set   = set(existing_nodes)

    # Collect only the genuinely new nodes (preserve order, deduplicate)
    new_nodes: List[str] = []
    for path in new_paths:
        for node in path:
            if node not in existing_set and node not in new_nodes:
                new_nodes.append(node)

    # Build merged chain: existing nodes first, then new terminal nodes appended
    merged_chain = existing_nodes + new_nodes
    state.chain      = merged_chain
    state.action_node = action_node

    # Build all_edges: reuse previous edges + add edges for each new path
    existing_edges_set = {(s, t) for s, t in prev_state.all_edges}
    merged_edges: List[Tuple[str, str]] = list(prev_state.all_edges)

    for path in new_paths:
        for src, tgt in zip(path[:-1], path[1:]):
            if (src, tgt) not in existing_edges_set:
                merged_edges.append((src, tgt))
                existing_edges_set.add((src, tgt))

    state.all_edges = merged_edges
    state.paths     = (prev_state.paths or [list(prev_state.chain)]) + new_paths

    # Apply traversal-pattern guard to merged chain
    state.chain     = _rewrite_traversal_chain(state.task, state.chain)

    # --- RAG: reuse existing edge_apis, run retrieval only for new edges ---
    prev_edge_map: dict = {}
    for edge, api in zip(prev_state.all_edges, prev_state.edge_apis or []):
        prev_edge_map[edge] = api

    edge_apis  = []
    api_parts  = []
    for edge in state.all_edges:
        if edge in prev_edge_map:
            hit = prev_edge_map[edge]
            edge_apis.append(hit)
            if hit:
                method = hit["function_name"].split("(")[0].split(".")[-1].strip()
                api_parts.append(
                    f"{edge[0].split('.')[-1]}->{edge[1].split('.')[-1]}: {method} [reused]"
                )
                print(f"  [extend/rag] [{edge[0]}] -> [{edge[1]}]  reused: {hit['function_name']}",
                      flush=True)
            else:
                api_parts.append(
                    f"{edge[0].split('.')[-1]}->{edge[1].split('.')[-1]}: NOT FOUND [reused]"
                )
        else:
            hit = _rag_query_for_edge(edge[0], edge[1], embed_model, metadata, embeddings)
            edge_apis.append(hit)
            if hit:
                method = hit["function_name"].split("(")[0].split(".")[-1].strip()
                api_parts.append(
                    f"{edge[0].split('.')[-1]}->{edge[1].split('.')[-1]}: {method} [new]"
                )
                print(f"  [extend/rag] [{edge[0]}] -> [{edge[1]}]  NEW hit: {hit['function_name']}",
                      flush=True)
            else:
                api_parts.append(
                    f"{edge[0].split('.')[-1]}->{edge[1].split('.')[-1]}: NOT FOUND [new]"
                )
                print(f"  [extend/rag] [{edge[0]}] -> [{edge[1]}]  NEW hit: [none]", flush=True)

    state.edge_apis   = edge_apis
    state.api_summary = " | ".join(api_parts)
    state.add_bootstrap_obs(
        "causal_extend",
        f"extended chain: {' -> '.join(state.chain)}  |  edges: {state.api_summary}",
    )


# ─────────────────────────────────────────────────────────────────────────────
# Option 2 case runner — respects state._bootstrap_done
# ─────────────────────────────────────────────────────────────────────────────

def run_case_option2(
    task: str,
    state: CausalAgentState,
    api_key: str,
    model: str,
    embed_model,
    metadata: list,
    embeddings,
    static_verifier: OpenROADStaticVerifier,
    causal_verifier: CausalVerifier,
    code_pieces: Optional[list] = None,
    llm_verifier=None,
) -> CausalAgentState:
    """Bootstrap → controller loop → commit.

    Identical to run_case except: when state._bootstrap_done is True (chain was
    already extended by bootstrap_causal_extend), skip extract+RAG and jump
    straight to generate+verify.
    """
    print(f"\n{'─'*60}", flush=True)
    print(f"TASK: {task}", flush=True)
    print(f"BUDGET: {state.max_budget} controller actions", flush=True)

    if getattr(state, "_bootstrap_done", False):
        # Chain + edge_apis already set by bootstrap_causal_extend — skip extract+RAG
        print(f"[Pipeline] [extended] generate → causal_verify (chain reused)", flush=True)
        bootstrap_generate(state, api_key, model)
        bootstrap_static_verify(state, static_verifier, causal_verifier)
        print(f"[Bootstrap] done — committing generated code for OpenROAD execution\n",
              flush=True)
    else:
        run_bootstrap(state, api_key, model, embed_model, metadata, embeddings,
                      static_verifier, causal_verifier)

    # ── same commit / controller logic as run_case ─────────────────────────────
    if state.static_result and state.static_result.passed:
        if llm_verifier is not None and state.current_code:
            from dispatcher import _build_chain_context
            chain_ctx = _build_chain_context(state.chain, state.edge_apis)
            lv_snap   = llm_verifier.verify(task, state.current_code, chain_ctx)
            state.llm_result = lv_snap
            if lv_snap.passed:
                print(f"  [loop] bootstrap PASS + LLM PASS (conf={lv_snap.confidence:.2f})"
                      f" — committing directly", flush=True)
                state.committed_code = state.best_code or state.current_code
                state.committed      = True
                return state
            else:
                print(f"  [loop] bootstrap static PASS but LLM FAIL(L5) "
                      f"conf={lv_snap.confidence:.2f} — entering controller loop",
                      flush=True)
        else:
            print(f"  [loop] bootstrap passed — committing directly", flush=True)
            state.committed_code = state.best_code or state.current_code
            state.committed      = True
            return state

    ctrl       = CausalController(api_key=api_key, model=model)
    dispatcher = CausalDispatcher(
        api_key=api_key, model=model,
        embed_model=embed_model, metadata=metadata, embeddings=embeddings,
        static_verifier=static_verifier, causal_verifier=causal_verifier,
        code_pieces=code_pieces or [],
        llm_verifier=llm_verifier,
    )
    dispatcher.reset_conversation(state)

    while not state.committed and state.budget_remaining > 0:
        action      = ctrl.decide(state)
        observation = dispatcher.execute(action, state)
        state.add_observation(action.next_action, observation, action.diagnosis or "")

    if not state.committed:
        state.committed_code = state.best_code or state.current_code
        state.committed      = True
        print(f"  [loop] budget exhausted — committing best (score={state.best_score:.3f})",
              flush=True)

    return state


# ─────────────────────────────────────────────────────────────────────────────
# Case loader  (identical logic to run_agent_sequential.py)
# ─────────────────────────────────────────────────────────────────────────────

def load_cases(path: str, level: int) -> List[Tuple[str, List[str]]]:
    """Return list of (complex_prompt, steps[]) from a Format-B dataset.

    Expected column layout (one row per case):
        complex_prompt | step_1 | step_2 | step_3 | ... | step_N

    level=N reads columns 1..N (step_1 through step_N).
    Rows with an empty complex_prompt are skipped.
    Steps beyond available columns are returned as empty strings.
    """
    import openpyxl
    wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    cases = []
    for row in rows[1:]:                          # skip header row
        if not row[0] or not str(row[0]).strip():
            continue
        complex_prompt = str(row[0]).strip()
        steps = []
        for col_idx in range(1, level + 1):       # cols 1..level
            val = row[col_idx] if len(row) > col_idx else None
            steps.append(str(val).strip() if val and str(val).strip() else "")
        cases.append((complex_prompt, steps))
    return cases


# ─────────────────────────────────────────────────────────────────────────────
# OpenROAD execution helper (runs in an already-open process)
# ─────────────────────────────────────────────────────────────────────────────

def _send_code(proc, code: str, oq: queue.Queue,
               max_wait: int, flush_time: float) -> Tuple[str, bool]:
    cmd = processCodeString(code)
    return sendCommandOpenROAD(proc, cmd, oq, max_wait, flush_time)


# ─────────────────────────────────────────────────────────────────────────────
# Main eval loop
# ─────────────────────────────────────────────────────────────────────────────

def Run(
    action_db_path:     str,
    result_path:        str,
    rag_api_path:       str,
    rag_code_path:      str,
    openroad_path:      str,
    api_key:            str,
    openai_model:       str           = "gpt-4.1-mini",
    level:              int           = 2,
    budget:             int           = 6,
    load_design_time:   int           = 5,
    max_wait_time:      int           = 120,
    command_flush_time: float         = 0.1,
    num_cases:          Optional[int] = None,
    start_case:         Optional[int] = None,
    cases:              Optional[List[int]] = None,
    run_dir:            str           = "",
    run_tag:            str           = "",
):
    # ── RAG setup ─────────────────────────────────────────────────────────────
    print("\nLoading embedding model...", flush=True)
    embed_model = SentenceTransformer("all-MiniLM-L6-v2")

    rag_df    = pd.read_csv(rag_api_path)
    metadata  = []
    documents = []
    for _, row in rag_df.iterrows():
        desc = str(row.get("Description:", "")).strip()
        if not desc or desc.lower() == "nan":
            continue
        documents.append(f"OpenROAD Python API Description:{desc}")
        metadata.append(row.to_dict())
    print(f"  [RAG] Encoding {len(documents)} API entries...", flush=True)
    embeddings = embed_model.encode(documents, convert_to_tensor=True,
                                    show_progress_bar=False)

    # ── verifiers ─────────────────────────────────────────────────────────────
    static_ver = OpenROADStaticVerifier(rag_api_path)
    causal_ver = CausalVerifier(metadata=metadata)
    llm_ver    = CausalLLMVerifier(api_key=api_key, model=openai_model, fail_open=True)
    print("[Init] Static + causal + LLM semantic verifiers ready.", flush=True)

    # ── code examples (RAGCodePiece.csv) ──────────────────────────────────────
    code_pieces = []
    if rag_code_path and os.path.isfile(rag_code_path):
        cp_df = pd.read_csv(rag_code_path)
        for _, row in cp_df.iterrows():
            desc = str(row.get("Description:", "")).strip()
            code = str(row.get("Code Piece:", "")).strip()
            if desc and code and desc.lower() != "nan" and code.lower() != "nan":
                code_pieces.append({"description": desc, "code": code})
        print(f"[Init] Loaded {len(code_pieces)} code examples.", flush=True)
    else:
        print("[Init] RAGCodePiecePath not set — code examples disabled.", flush=True)

    # ── OpenROAD PTY ──────────────────────────────────────────────────────────
    master_fd, slave_fd = os.openpty()
    oq                  = queue.Queue()
    threading.Thread(
        target=readOpenROADOutput,
        args=(master_fd, oq, "STDOUT", threading.Event()),
        daemon=True,
    ).start()
    print("[Init] OpenROAD PTY ready.", flush=True)

    # ── load + slice cases ────────────────────────────────────────────────────
    all_cases = load_cases(action_db_path, level)
    if cases is not None:
        indexed_cases = [(i, all_cases[i][0], all_cases[i][1])
                         for i in cases if i < len(all_cases)]
    else:
        if start_case is not None:
            all_cases = all_cases[start_case:]
        if num_cases is not None:
            all_cases = all_cases[:num_cases]
        offset = start_case or 0
        indexed_cases = [(offset + i, cp, steps)
                         for i, (cp, steps) in enumerate(all_cases)]
    print(f"[Init] Level={level}, running {len(indexed_cases)} case(s).", flush=True)

    # ── result workbook ───────────────────────────────────────────────────────
    os.makedirs(result_path, exist_ok=True)
    dataset_tag  = os.path.splitext(os.path.basename(action_db_path))[0]
    tag_suffix   = f"_{run_tag}" if run_tag else ""
    _result_file = os.path.join(
        result_path,
        f"{openai_model.replace('.', '-')}__causal_L{level}_{dataset_tag}_opt2{tag_suffix}.xlsx"
    )

    if os.path.exists(_result_file):
        wb = load_workbook(_result_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        # col 1 = complex_prompt; stepN_prompt/code/result/output; overall; lessons
        headers = {1: "complex_prompt"}
        col = 2
        for n in range(1, level + 1):
            for field in ("prompt", "code", "causal_verdict", "result", "output"):
                headers[col] = f"step{n}_{field}"
                col += 1
        headers[col] = "overall";    col += 1
        for n in range(1, level + 1):
            headers[col] = f"step{n}_lessons"; col += 1
        for c, h in headers.items():
            ws.cell(row=1, column=c, value=h)

    # Column index helpers
    _FIELDS = ["prompt", "code", "causal_verdict", "result", "output"]

    def _col(step_n: int, field: str) -> int:
        return (step_n - 1) * len(_FIELDS) + _FIELDS.index(field) + 2

    def _overall_col() -> int:
        return level * len(_FIELDS) + 2

    def _lesson_col(step_n: int) -> int:
        return level * len(_FIELDS) + 2 + step_n

    def _clean(s):
        return ILLEGAL_CHARACTERS_RE.sub("", str(s or ""))

    passed_all = 0
    total      = 0

    # ── case loop ─────────────────────────────────────────────────────────────
    for case_i, complex_prompt, steps in indexed_cases:
        row   = case_i + 2
        total += 1
        print(f"\n{'='*60}", flush=True)
        print(f"CASE {case_i+1}/{len(indexed_cases)}  (level={level}, {len(steps)} steps)",
              flush=True)

        # Fresh OpenROAD process for this case
        orig_dir = os.getcwd()
        if run_dir:
            os.chdir(run_dir)
        proc = runOpenROADShell(openroad_path, load_design_time, slave_fd, "")
        if run_dir:
            os.chdir(orig_dir)
        time.sleep(load_design_time)
        clearQueue(oq)

        accepted_codes: List[str]          = []
        step_states:    List[CausalAgentState] = []
        step_verdicts:  List[str]          = []   # causal verifier verdict string
        step_results:   List[bool]         = []   # True = OpenROAD PASS
        step_outputs:   List[str]          = []

        for step_idx, raw_prompt in enumerate(steps):
            step_n = step_idx + 1
            print(f"\n  ── Step {step_n}/{len(steps)}: {raw_prompt[:100]}", flush=True)

            # Skip if any prior step failed in OpenROAD
            if step_idx > 0 and not step_results[-1]:
                print(f"  [Step {step_n}] SKIP — step {step_n-1} failed", flush=True)
                step_states.append(None)
                step_verdicts.append("SKIP")
                step_results.append(False)
                step_outputs.append("")
                continue

            # Build task: prepend prior accepted codes as live-session context
            if accepted_codes:
                ctx_block = "\n\n".join(
                    f"# --- Step {i+1} (already executed, do NOT re-run) ---\n{c}"
                    for i, c in enumerate(accepted_codes)
                )
                task = (
                    f"{raw_prompt}\n\n"
                    f"[Context] The following code already ran successfully in this "
                    f"OpenROAD session. The objects it created are live in memory — "
                    f"build on them directly, do NOT re-execute this code.\n"
                    f"```python\n{ctx_block}\n```"
                )
            else:
                task = raw_prompt

            # ── Option 2: fresh extract for step 1, extend chain for step 2+ ──
            state = CausalAgentState(task=task, max_budget=budget)

            if step_idx == 0 or not step_states:
                # Step 1: normal bootstrap (extract chain fresh)
                print(f"  [Option2] Step 1 — fresh chain extraction", flush=True)
                # run_case calls bootstrap internally
            else:
                # Step N>1: extend the previous step's chain instead of re-extracting
                prev_state = next(
                    (s for s in reversed(step_states) if s is not None), None
                )
                if prev_state is not None:
                    print(
                        f"  [Option2] Step {step_n} — extending chain from step "
                        f"{step_states.index(prev_state)+1}: "
                        f"{' -> '.join(prev_state.chain)}",
                        flush=True,
                    )
                    bootstrap_causal_extend(
                        state, prev_state,
                        api_key=api_key, model=openai_model,
                        embed_model=embed_model, metadata=metadata, embeddings=embeddings,
                    )
                    # Mark bootstrap as already done so run_case skips re-extraction
                    state._bootstrap_done = True
                else:
                    print(f"  [Option2] Step {step_n} — no prior state; falling back to fresh",
                          flush=True)

            state = run_case_option2(
                task, state,
                api_key=api_key, model=openai_model,
                embed_model=embed_model, metadata=metadata, embeddings=embeddings,
                static_verifier=static_ver, causal_verifier=causal_ver,
                code_pieces=code_pieces,
                llm_verifier=llm_ver,
            )
            step_states.append(state)

            # causal verifier verdict string
            sv = state.static_result
            if sv:
                causal_verdict = (
                    "PASS" if sv.passed
                    else f"FAIL(L{sv.layer_failed}): {'; '.join(sv.issues[:2])}"
                )
            else:
                causal_verdict = "N/A"
            step_verdicts.append(causal_verdict)

            # ── run in SAME OpenROAD process ───────────────────────────────
            stdout, has_tb = "", True
            if state.committed_code:
                try:
                    stdout, has_tb = _send_code(
                        proc, state.committed_code, oq,
                        max_wait_time, command_flush_time,
                    )
                except Exception as exc:
                    stdout = f"[OpenROAD error] {exc}"
                    has_tb = True

            passed = bool(state.committed_code) and not has_tb
            step_results.append(passed)
            step_outputs.append(stdout)

            if passed:
                accepted_codes.append(state.committed_code)
            print(f"  [Step {step_n} OpenROAD] {'PASS' if passed else 'FAIL'}", flush=True)
            if not passed and stdout:
                print(f"  [Step {step_n} error]\n{stdout.strip()}", flush=True)

        # terminate OpenROAD after all steps
        try:
            proc.terminate()
            proc.wait()
        except Exception:
            pass

        all_passed = all(step_results)
        if all_passed:
            passed_all += 1

        overall = "PASS" if all_passed else (
            f"STEP{next(i for i, r in enumerate(step_results) if not r)+1}_FAIL"
        )
        print(f"  [Overall] {overall}", flush=True)

        # ── write Excel ────────────────────────────────────────────────────
        ws.cell(row=row, column=1, value=_clean(complex_prompt))

        for step_idx, (raw_prompt, state, verdict, passed, stdout) in enumerate(
            zip(steps, step_states, step_verdicts, step_results, step_outputs)
        ):
            step_n = step_idx + 1
            ws.cell(row=row, column=_col(step_n, "prompt"),
                    value=raw_prompt)
            ws.cell(row=row, column=_col(step_n, "code"),
                    value=_clean(state.committed_code if state else ""))
            ws.cell(row=row, column=_col(step_n, "causal_verdict"),
                    value=verdict)
            ws.cell(row=row, column=_col(step_n, "result"),
                    value="PASS" if passed else ("SKIP" if state is None else "FAIL"))
            ws.cell(row=row, column=_col(step_n, "output"),
                    value=_clean(stdout[:2000]))
            ws.cell(row=row, column=_lesson_col(step_n),
                    value=" | ".join(state.lessons) if state else "")

        ws.cell(row=row, column=_overall_col(), value=overall)
        wb.save(_result_file)

    print(f"\n{'='*60}", flush=True)
    print(f"All steps passed: {passed_all}/{total} = "
          f"{passed_all/max(total,1)*100:.1f}%", flush=True)
    print(f"Results → {_result_file}", flush=True)


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(
        description="Sequential N-step causal agentic eval from action_db.xlsx"
    )
    p.add_argument("--openaiKey",        required=True)
    p.add_argument("--model",            default="gpt-4.1-mini")
    p.add_argument("--actionDBPath",     default="src_1_reflector/action_db.xlsx")
    p.add_argument("--resultPath",       default="result/")
    p.add_argument("--RAGApiPath",       default="RAGData/RAGAPIs.csv")
    p.add_argument("--RAGCodePiecePath", default="RAGData/RAGCodePiece.csv")
    p.add_argument("--openroadPath",     default="OpenROAD/build/src/openroad")
    p.add_argument("--runDir",           default="src_1_reflector",
                   help="Working dir for OpenROAD shell")
    p.add_argument("--level",            type=int, default=2,
                   help="Number of steps to run (1..N)")
    p.add_argument("--budget",           type=int, default=6,
                   help="Causal controller action budget per step")
    p.add_argument("--loadDesignTime",   type=int, default=5)
    p.add_argument("--maxWaitTime",      type=int, default=120)
    p.add_argument("--commandFlushTime", type=float, default=0.1)
    p.add_argument("--numCases",         type=int, default=None)
    p.add_argument("--startCase",        type=int, default=None,
                   help="0-indexed case to start from")
    p.add_argument("--cases",            type=int, nargs="+", default=None,
                   help="Specific 0-indexed case numbers to run (e.g. --cases 2 6)")
    p.add_argument("--runTag",           default="",
                   help="Optional tag appended to result filename to avoid overwriting")
    args = p.parse_args()

    def _abs(s):
        return s if os.path.isabs(s) else os.path.join(_ROOT, s)

    Run(
        action_db_path     = _abs(args.actionDBPath),
        result_path        = _abs(args.resultPath),
        rag_api_path       = _abs(args.RAGApiPath),
        rag_code_path      = _abs(args.RAGCodePiecePath),
        openroad_path      = _abs(args.openroadPath),
        api_key            = args.openaiKey,
        openai_model       = args.model,
        level              = args.level,
        budget             = args.budget,
        load_design_time   = args.loadDesignTime,
        max_wait_time      = args.maxWaitTime,
        command_flush_time = args.commandFlushTime,
        num_cases          = args.numCases,
        start_case         = args.startCase,
        cases              = args.cases,
        run_dir            = _abs(args.runDir) if args.runDir else "",
        run_tag            = args.runTag,
    )


if __name__ == "__main__":
    main()
