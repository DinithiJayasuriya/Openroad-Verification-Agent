"""run_causal_agent_sequential.py — Sequential N-step causal agentic eval from action_db.xlsx.

Mirrors the flow of src_1_reflector/run_agent_sequential.py but uses the
causal pipeline (chain extraction → node-specific RAG → generation →
static+causal+LLM verification) instead of the RAG-based pipeline.

Controlled by --level:
  level 1 — runs only step 1 (level_one column)
  level 2 — runs step 1 → step 2 (level_one + level_two columns)
  level 3 — runs all 3 sub-prompts (level_one + level_two + level_three columns)

Each step runs in the SAME OpenROAD session (state carries over).
A step is only attempted if all prior steps passed in OpenROAD.
Each step's causal agent receives all previously accepted code as context.

Usage:
  python causal_verifier/run_causal_agent_sequential.py \\
    --openaiKey sk-... \\
    --actionDBPath src_1_reflector/action_db.xlsx \\
    --RAGApiPath RAGData/RAGAPIs.csv \\
    --RAGCodePiecePath RAGData/RAGCodePiece.csv \\
    --openroadPath OpenROAD/build/src/openroad \\
    --runDir src_1_reflector \\
    --resultPath result/ \\
    --level 3 --budget 6 --numCases 5
"""

import argparse
import os
import queue
import re
import sys
import threading
import time
from typing import List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from sentence_transformers import SentenceTransformer

# ── path setup ────────────────────────────────────────────────────────────────
_ROOT       = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_CAUSAL_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(_ROOT, "src_1_reflector"))
sys.path.insert(0, os.path.join(_ROOT, "src_1_agentic"))
sys.path.insert(0, _CAUSAL_DIR)

from util import (
    runOpenROADShell, sendCommandOpenROAD, processCodeString,
    clearQueue, readOpenROADOutput,
)
from verifier import OpenROADStaticVerifier

from causal_state    import CausalAgentState
from causal_verifier import CausalVerifier
from llm_verifier    import CausalLLMVerifier

# Re-use run_case and helpers from run_causal_agent
from run_causal_agent import run_case, _clean_output


# ─────────────────────────────────────────────────────────────────────────────
# Case loader  (identical logic to run_agent_sequential.py)
# ─────────────────────────────────────────────────────────────────────────────

def load_cases(path: str, level: int) -> List[Tuple[str, List[str]]]:
    """Return list of (complex_prompt, steps[]) from a Format-B dataset.

    Expected column layout (one row per case):
        complex_prompt | step_1 | step_2 | step_3 | ... | step_N

    level=N reads columns 1..N (step_1 through step_N).
    Rows with an empty complex_prompt are skipped.
    Steps beyond available columns are returned as empty strings.
    """
    import openpyxl
    wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    cases = []
    for row in rows[1:]:                          # skip header row
        if not row[0] or not str(row[0]).strip():
            continue
        complex_prompt = str(row[0]).strip()
        steps = []
        for col_idx in range(1, level + 1):       # cols 1..level
            val = row[col_idx] if len(row) > col_idx else None
            steps.append(str(val).strip() if val and str(val).strip() else "")
        cases.append((complex_prompt, steps))
    return cases


# ─────────────────────────────────────────────────────────────────────────────
# OpenROAD execution helper (runs in an already-open process)
# ─────────────────────────────────────────────────────────────────────────────

def _send_code(proc, code: str, oq: queue.Queue,
               max_wait: int, flush_time: float) -> Tuple[str, bool]:
    cmd = processCodeString(code)
    return sendCommandOpenROAD(proc, cmd, oq, max_wait, flush_time)


# ─────────────────────────────────────────────────────────────────────────────
# Main eval loop
# ─────────────────────────────────────────────────────────────────────────────

def Run(
    action_db_path:     str,
    result_path:        str,
    rag_api_path:       str,
    rag_code_path:      str,
    openroad_path:      str,
    api_key:            str,
    openai_model:       str           = "gpt-4.1-mini",
    level:              int           = 2,
    budget:             int           = 6,
    load_design_time:   int           = 5,
    max_wait_time:      int           = 120,
    command_flush_time: float         = 0.1,
    num_cases:          Optional[int] = None,
    start_case:         Optional[int] = None,
    cases:              Optional[List[int]] = None,
    run_dir:            str           = "",
    run_tag:            str           = "",
):
    # ── RAG setup ─────────────────────────────────────────────────────────────
    print("\nLoading embedding model...", flush=True)
    embed_model = SentenceTransformer("all-MiniLM-L6-v2")

    rag_df    = pd.read_csv(rag_api_path)
    metadata  = []
    documents = []
    for _, row in rag_df.iterrows():
        desc = str(row.get("Description:", "")).strip()
        if not desc or desc.lower() == "nan":
            continue
        documents.append(f"OpenROAD Python API Description:{desc}")
        metadata.append(row.to_dict())
    print(f"  [RAG] Encoding {len(documents)} API entries...", flush=True)
    embeddings = embed_model.encode(documents, convert_to_tensor=True,
                                    show_progress_bar=False)

    # ── verifiers ─────────────────────────────────────────────────────────────
    static_ver = OpenROADStaticVerifier(rag_api_path)
    causal_ver = CausalVerifier(metadata=metadata)
    llm_ver    = CausalLLMVerifier(api_key=api_key, model=openai_model, fail_open=True)
    print("[Init] Static + causal + LLM semantic verifiers ready.", flush=True)

    # ── code examples (RAGCodePiece.csv) ──────────────────────────────────────
    code_pieces = []
    if rag_code_path and os.path.isfile(rag_code_path):
        cp_df = pd.read_csv(rag_code_path)
        for _, row in cp_df.iterrows():
            desc = str(row.get("Description:", "")).strip()
            code = str(row.get("Code Piece:", "")).strip()
            if desc and code and desc.lower() != "nan" and code.lower() != "nan":
                code_pieces.append({"description": desc, "code": code})
        print(f"[Init] Loaded {len(code_pieces)} code examples.", flush=True)
    else:
        print("[Init] RAGCodePiecePath not set — code examples disabled.", flush=True)

    # ── OpenROAD PTY ──────────────────────────────────────────────────────────
    master_fd, slave_fd = os.openpty()
    oq                  = queue.Queue()
    threading.Thread(
        target=readOpenROADOutput,
        args=(master_fd, oq, "STDOUT", threading.Event()),
        daemon=True,
    ).start()
    print("[Init] OpenROAD PTY ready.", flush=True)

    # ── load + slice cases ────────────────────────────────────────────────────
    all_cases = load_cases(action_db_path, level)
    if cases is not None:
        indexed_cases = [(i, all_cases[i][0], all_cases[i][1])
                         for i in cases if i < len(all_cases)]
    else:
        if start_case is not None:
            all_cases = all_cases[start_case:]
        if num_cases is not None:
            all_cases = all_cases[:num_cases]
        offset = start_case or 0
        indexed_cases = [(offset + i, cp, steps)
                         for i, (cp, steps) in enumerate(all_cases)]
    print(f"[Init] Level={level}, running {len(indexed_cases)} case(s).", flush=True)

    # ── result workbook ───────────────────────────────────────────────────────
    os.makedirs(result_path, exist_ok=True)
    dataset_tag  = os.path.splitext(os.path.basename(action_db_path))[0]
    tag_suffix   = f"_{run_tag}" if run_tag else ""
    _result_file = os.path.join(
        result_path,
        f"{openai_model.replace('.', '-')}__causal_L{level}_{dataset_tag}{tag_suffix}.xlsx"
    )

    if os.path.exists(_result_file):
        wb = load_workbook(_result_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        # col 1 = complex_prompt; stepN_prompt/code/result/output; overall; lessons
        headers = {1: "complex_prompt"}
        col = 2
        for n in range(1, level + 1):
            for field in ("prompt", "code", "causal_verdict", "result", "output"):
                headers[col] = f"step{n}_{field}"
                col += 1
        headers[col] = "overall";    col += 1
        for n in range(1, level + 1):
            headers[col] = f"step{n}_lessons"; col += 1
        for c, h in headers.items():
            ws.cell(row=1, column=c, value=h)

    # Column index helpers
    _FIELDS = ["prompt", "code", "causal_verdict", "result", "output"]

    def _col(step_n: int, field: str) -> int:
        return (step_n - 1) * len(_FIELDS) + _FIELDS.index(field) + 2

    def _overall_col() -> int:
        return level * len(_FIELDS) + 2

    def _lesson_col(step_n: int) -> int:
        return level * len(_FIELDS) + 2 + step_n

    def _clean(s):
        return ILLEGAL_CHARACTERS_RE.sub("", str(s or ""))

    passed_all = 0
    total      = 0

    # ── case loop ─────────────────────────────────────────────────────────────
    for case_i, complex_prompt, steps in indexed_cases:
        row   = case_i + 2
        total += 1
        print(f"\n{'='*60}", flush=True)
        print(f"CASE {case_i+1}/{len(indexed_cases)}  (level={level}, {len(steps)} steps)",
              flush=True)

        # Fresh OpenROAD process for this case
        orig_dir = os.getcwd()
        if run_dir:
            os.chdir(run_dir)
        proc = runOpenROADShell(openroad_path, load_design_time, slave_fd, "")
        if run_dir:
            os.chdir(orig_dir)
        time.sleep(load_design_time)
        clearQueue(oq)

        accepted_codes: List[str]          = []
        step_states:    List[CausalAgentState] = []
        step_verdicts:  List[str]          = []   # causal verifier verdict string
        step_results:   List[bool]         = []   # True = OpenROAD PASS
        step_outputs:   List[str]          = []

        for step_idx, raw_prompt in enumerate(steps):
            step_n = step_idx + 1
            print(f"\n  ── Step {step_n}/{len(steps)}: {raw_prompt[:100]}", flush=True)

            # Skip if any prior step failed in OpenROAD
            if step_idx > 0 and not step_results[-1]:
                print(f"  [Step {step_n}] SKIP — step {step_n-1} failed", flush=True)
                step_states.append(None)
                step_verdicts.append("SKIP")
                step_results.append(False)
                step_outputs.append("")
                continue

            # Build task: prepend prior accepted codes as live-session context
            if accepted_codes:
                ctx_block = "\n\n".join(
                    f"# --- Step {i+1} (already executed, do NOT re-run) ---\n{c}"
                    for i, c in enumerate(accepted_codes)
                )
                task = (
                    f"{raw_prompt}\n\n"
                    f"[Context] The following code already ran successfully in this "
                    f"OpenROAD session. The objects it created are live in memory — "
                    f"build on them directly, do NOT re-execute this code.\n"
                    f"```python\n{ctx_block}\n```"
                )
            else:
                task = raw_prompt

            # ── causal agent loop ──────────────────────────────────────────
            state = CausalAgentState(task=task, max_budget=budget)
            state = run_case(
                task, state,
                api_key=api_key, model=openai_model,
                embed_model=embed_model, metadata=metadata, embeddings=embeddings,
                static_verifier=static_ver, causal_verifier=causal_ver,
                code_pieces=code_pieces,
                llm_verifier=llm_ver,
            )
            step_states.append(state)

            # causal verifier verdict string
            sv = state.static_result
            if sv:
                causal_verdict = (
                    "PASS" if sv.passed
                    else f"FAIL(L{sv.layer_failed}): {'; '.join(sv.issues[:2])}"
                )
            else:
                causal_verdict = "N/A"
            step_verdicts.append(causal_verdict)

            # ── run in SAME OpenROAD process ───────────────────────────────
            stdout, has_tb = "", True
            if state.committed_code:
                try:
                    stdout, has_tb = _send_code(
                        proc, state.committed_code, oq,
                        max_wait_time, command_flush_time,
                    )
                except Exception as exc:
                    stdout = f"[OpenROAD error] {exc}"
                    has_tb = True

            passed = bool(state.committed_code) and not has_tb
            step_results.append(passed)
            step_outputs.append(stdout)

            if passed:
                accepted_codes.append(state.committed_code)
            print(f"  [Step {step_n} OpenROAD] {'PASS' if passed else 'FAIL'}", flush=True)
            if not passed and stdout:
                print(f"  [Step {step_n} error]\n{stdout.strip()}", flush=True)

        # terminate OpenROAD after all steps
        try:
            proc.terminate()
            proc.wait()
        except Exception:
            pass

        all_passed = all(step_results)
        if all_passed:
            passed_all += 1

        overall = "PASS" if all_passed else (
            f"STEP{next(i for i, r in enumerate(step_results) if not r)+1}_FAIL"
        )
        print(f"  [Overall] {overall}", flush=True)

        # ── write Excel ────────────────────────────────────────────────────
        ws.cell(row=row, column=1, value=_clean(complex_prompt))

        for step_idx, (raw_prompt, state, verdict, passed, stdout) in enumerate(
            zip(steps, step_states, step_verdicts, step_results, step_outputs)
        ):
            step_n = step_idx + 1
            ws.cell(row=row, column=_col(step_n, "prompt"),
                    value=raw_prompt)
            ws.cell(row=row, column=_col(step_n, "code"),
                    value=_clean(state.committed_code if state else ""))
            ws.cell(row=row, column=_col(step_n, "causal_verdict"),
                    value=verdict)
            ws.cell(row=row, column=_col(step_n, "result"),
                    value="PASS" if passed else ("SKIP" if state is None else "FAIL"))
            ws.cell(row=row, column=_col(step_n, "output"),
                    value=_clean(stdout[:2000]))
            ws.cell(row=row, column=_lesson_col(step_n),
                    value=" | ".join(state.lessons) if state else "")

        ws.cell(row=row, column=_overall_col(), value=overall)
        wb.save(_result_file)

    print(f"\n{'='*60}", flush=True)
    print(f"All steps passed: {passed_all}/{total} = "
          f"{passed_all/max(total,1)*100:.1f}%", flush=True)
    print(f"Results → {_result_file}", flush=True)


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(
        description="Sequential N-step causal agentic eval from action_db.xlsx"
    )
    p.add_argument("--openaiKey",        required=True)
    p.add_argument("--model",            default="gpt-4.1-mini")
    p.add_argument("--actionDBPath",     default="src_1_reflector/action_db.xlsx")
    p.add_argument("--resultPath",       default="result/")
    p.add_argument("--RAGApiPath",       default="RAGData/RAGAPIs.csv")
    p.add_argument("--RAGCodePiecePath", default="RAGData/RAGCodePiece.csv")
    p.add_argument("--openroadPath",     default="OpenROAD/build/src/openroad")
    p.add_argument("--runDir",           default="src_1_reflector",
                   help="Working dir for OpenROAD shell")
    p.add_argument("--level",            type=int, default=2,
                   help="Number of steps to run (1..N)")
    p.add_argument("--budget",           type=int, default=6,
                   help="Causal controller action budget per step")
    p.add_argument("--loadDesignTime",   type=int, default=5)
    p.add_argument("--maxWaitTime",      type=int, default=120)
    p.add_argument("--commandFlushTime", type=float, default=0.1)
    p.add_argument("--numCases",         type=int, default=None)
    p.add_argument("--startCase",        type=int, default=None,
                   help="0-indexed case to start from")
    p.add_argument("--cases",            type=int, nargs="+", default=None,
                   help="Specific 0-indexed case numbers to run (e.g. --cases 2 6)")
    p.add_argument("--runTag",           default="",
                   help="Optional tag appended to result filename to avoid overwriting")
    args = p.parse_args()

    def _abs(s):
        return s if os.path.isabs(s) else os.path.join(_ROOT, s)

    Run(
        action_db_path     = _abs(args.actionDBPath),
        result_path        = _abs(args.resultPath),
        rag_api_path       = _abs(args.RAGApiPath),
        rag_code_path      = _abs(args.RAGCodePiecePath),
        openroad_path      = _abs(args.openroadPath),
        api_key            = args.openaiKey,
        openai_model       = args.model,
        level              = args.level,
        budget             = args.budget,
        load_design_time   = args.loadDesignTime,
        max_wait_time      = args.maxWaitTime,
        command_flush_time = args.commandFlushTime,
        num_cases          = args.numCases,
        start_case         = args.startCase,
        cases              = args.cases,
        run_dir            = _abs(args.runDir) if args.runDir else "",
        run_tag            = args.runTag,
    )


if __name__ == "__main__":
    main()
