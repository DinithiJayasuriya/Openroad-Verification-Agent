"""run_baseline_llm_rag.py — Baseline: single-shot LLM + RAG (no verifier, no causal chain).

For each multi-step case in the stage_specific_tasks dataset:
  1. Retrieve top-k RAG docs for the step prompt (semantic similarity)
  2. Send task + RAG context to GPT → get code (single shot, no retry)
  3. Execute in OpenROAD
  4. Record pass/fail

This is the simplest possible baseline to compare against the full
causal pipeline.

Usage:
  python causal_verifier_4_2/run_baseline_llm_rag.py \
    --openaiKey sk-... \
    --actionDBPath causal_verifier_4_5_developed/stage_specific_tasks.xlsx \
    --RAGApiPath RAGData/RAGAPIs.csv \
    --RAGCodePiecePath RAGData/RAGCodePiece.csv \
    --openroadPath OpenROAD/build/src/openroad \
    --runDir src_1_reflector \
    --level 4 \
    --resultPath result/stage_ablation/
"""

import argparse
import json
import os
import queue
import re
import sys
import threading
import time
import urllib.request
import urllib.error
from typing import List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from sentence_transformers import SentenceTransformer
from sentence_transformers.util import cos_sim

# ── path setup ────────────────────────────────────────────────────────────────
_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(_ROOT, "src_1_reflector"))

from util import (
    runOpenROADShell, sendCommandOpenROAD, processCodeString,
    clearQueue, readOpenROADOutput,
)


# ── Simple system prompt (no causal chain, no verifier references) ────────────
_SYSTEM_PROMPT = """\
You are an expert OpenROAD Python API programmer.
Generate a Python script that runs inside the OpenROAD interactive Python shell.

Rules you MUST follow:
1. `design` (openroad.Design) and `tech` (openroad.Tech) are already available.
2. ALL methods are INSTANCE methods — call them on objects, never on the class.
3. Write FLAT procedural code — no functions, no classes, no `global`.
4. Print all results so they are visible in the shell output.
5. After any lookup that can return None (findInst, findNet, findBTerm, findITerm, etc.), \
immediately check if the result is None and print a clear message, then skip remaining steps \
using an if/else block. Never call exit() or sys.exit().

Output ONLY the Python code — no markdown fences, no explanation.
"""


# ── OpenAI call ───────────────────────────────────────────────────────────────
def _call_openai(messages, api_key, model, max_tokens=800) -> str:
    payload = json.dumps({
        "model": model, "messages": messages,
        "temperature": 0, "max_tokens": max_tokens,
    }).encode()
    req = urllib.request.Request(
        "https://api.openai.com/v1/chat/completions",
        data=payload,
        headers={"Content-Type": "application/json",
                 "Authorization": f"Bearer {api_key}"},
        method="POST",
    )
    wait = 10
    for _ in range(5):
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                body = json.loads(resp.read().decode())
            return body["choices"][0]["message"]["content"].strip()
        except urllib.error.HTTPError as e:
            if e.code == 429:
                print(f"  [rate-limit] waiting {wait}s...", flush=True)
                time.sleep(wait); wait = min(wait * 2, 120)
            else:
                print(f"  [API error] {e.code}", flush=True)
                return ""
        except Exception as e:
            print(f"  [API error] {e}", flush=True)
            return ""
    return ""


# ── RAG retrieval (vanilla semantic search) ───────────────────────────────────
def retrieve_rag(query: str, embeddings, embed_model, documents, metadata,
                 top_k: int = 10, threshold: float = 0.5) -> str:
    q_emb = embed_model.encode(query, convert_to_tensor=True)
    scores = cos_sim(q_emb, embeddings).cpu().numpy().flatten()
    top_idx = np.argsort(scores)[-top_k:][::-1]
    docs = []
    for idx in top_idx:
        if scores[idx] < threshold:
            break
        entry = metadata[idx]
        docs.append("\n".join(f"# {k} {v}" for k, v in entry.items()))
    if not docs:
        return ""
    return "\n\n".join(docs)


# ── Code extraction ──────────────────────────────────────────────────────────
def _extract_code(text: str) -> str:
    # Strip markdown fences if present
    if "```" in text:
        lines = []
        in_block = False
        for ln in text.splitlines():
            if ln.strip().startswith("```"):
                in_block = not in_block
                continue
            if in_block or not any(text.count("```") > 1 for _ in [0]):
                lines.append(ln)
        text = "\n".join(lines).strip()
        if not text:
            # fallback: just strip the fences
            text = re.sub(r"```\w*\n?", "", text).strip()

    # If [Code]: marker exists, take everything after it
    m = re.search(r"\[Code\]\s*:?\s*\n", text)
    if m:
        text = text[m.end():]

    return text.strip()


# ── Load cases from xlsx ─────────────────────────────────────────────────────
def load_cases(path: str, level: int, sheet: Optional[str] = None):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    cases = []
    for row in rows[1:]:
        if not row[0] or not str(row[0]).strip():
            continue
        complex_prompt = str(row[0]).strip()
        steps = []
        for col_idx in range(1, level + 1):
            val = row[col_idx] if len(row) > col_idx else None
            steps.append(str(val).strip() if val and str(val).strip() else "")
        cases.append((complex_prompt, steps))
    return cases


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    p = argparse.ArgumentParser(description="Baseline: LLM + RAG (no verifier, no causal chain)")
    p.add_argument("--openaiKey",        required=True)
    p.add_argument("--model",            default="gpt-4.1-mini")
    p.add_argument("--actionDBPath",     required=True)
    p.add_argument("--resultPath",       default="result/")
    p.add_argument("--RAGApiPath",       default="RAGData/RAGAPIs.csv")
    p.add_argument("--RAGCodePiecePath", default="RAGData/RAGCodePiece.csv")
    p.add_argument("--openroadPath",     default="OpenROAD/build/src/openroad")
    p.add_argument("--runDir",           default="src_1_reflector")
    p.add_argument("--level",            type=int, default=4)
    p.add_argument("--loadDesignTime",   type=int, default=5)
    p.add_argument("--maxWaitTime",      type=int, default=120)
    p.add_argument("--commandFlushTime", type=float, default=0.1)
    p.add_argument("--numCases",         type=int, default=None)
    p.add_argument("--startCase",        type=int, default=None)
    p.add_argument("--cases",            type=int, nargs="+", default=None)
    p.add_argument("--sheet",            default=None)
    p.add_argument("--runTag",           default="baseline_llm_rag")
    args = p.parse_args()

    def _abs(s):
        return s if os.path.isabs(s) else os.path.join(_ROOT, s)

    # ── RAG setup ─────────────────────────────────────────────────────────────
    print("\nLoading embedding model...", flush=True)
    embed_model = SentenceTransformer("all-MiniLM-L6-v2")

    rag_df = pd.read_csv(_abs(args.RAGApiPath))
    metadata = []
    documents = []
    for _, row in rag_df.iterrows():
        desc = str(row.get("Description:", "")).strip()
        if not desc or desc.lower() == "nan":
            continue
        documents.append(f"OpenROAD Python API Description:{desc}")
        metadata.append(row.to_dict())
    print(f"  [RAG] Encoding {len(documents)} API entries...", flush=True)
    embeddings = embed_model.encode(documents, convert_to_tensor=True,
                                    show_progress_bar=False)

    # ── Code examples ─────────────────────────────────────────────────────────
    code_examples_str = ""
    rag_code_path = _abs(args.RAGCodePiecePath)
    if os.path.isfile(rag_code_path):
        cp_df = pd.read_csv(rag_code_path)
        examples = []
        for _, row in cp_df.iterrows():
            desc = str(row.get("Description:", "")).strip()
            code = str(row.get("Code Piece:", "")).strip()
            if desc and code and desc.lower() != "nan" and code.lower() != "nan":
                examples.append(f"# Example: {desc}\n{code}")
        if examples:
            code_examples_str = "\n\n[Code Examples]\n" + "\n\n".join(examples[:5])
            print(f"  [RAG] Loaded {len(examples)} code examples.", flush=True)

    # ── OpenROAD PTY ──────────────────────────────────────────────────────────
    master_fd, slave_fd = os.openpty()
    oq = queue.Queue()
    threading.Thread(
        target=readOpenROADOutput,
        args=(master_fd, oq, "STDOUT", threading.Event()),
        daemon=True,
    ).start()

    # ── Load cases ────────────────────────────────────────────────────────────
    all_cases = load_cases(_abs(args.actionDBPath), args.level, sheet=args.sheet)
    if args.cases is not None:
        indexed_cases = [(i, all_cases[i][0], all_cases[i][1])
                         for i in args.cases if i < len(all_cases)]
    else:
        if args.startCase is not None:
            all_cases = all_cases[args.startCase:]
        if args.numCases is not None:
            all_cases = all_cases[:args.numCases]
        offset = args.startCase or 0
        indexed_cases = [(offset + i, cp, steps)
                         for i, (cp, steps) in enumerate(all_cases)]
    print(f"[Init] Level={args.level}, running {len(indexed_cases)} case(s).\n", flush=True)

    # ── Result workbook ───────────────────────────────────────────────────────
    os.makedirs(_abs(args.resultPath), exist_ok=True)
    dataset_tag = os.path.splitext(os.path.basename(args.actionDBPath))[0]
    result_file = os.path.join(
        _abs(args.resultPath),
        f"{args.model.replace('.', '-')}__baseline_L{args.level}_{dataset_tag}_{args.runTag}.xlsx"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    headers = ["complex_prompt"]
    for n in range(1, args.level + 1):
        for field in ("prompt", "rag_context", "code", "result", "output"):
            headers.append(f"step{n}_{field}")
    headers.append("overall")
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    _FIELDS = ["prompt", "rag_context", "code", "result", "output"]

    def _col(step_n, field):
        return (step_n - 1) * len(_FIELDS) + _FIELDS.index(field) + 2

    def _overall_col():
        return args.level * len(_FIELDS) + 2

    def _clean(s):
        return ILLEGAL_CHARACTERS_RE.sub("", str(s or ""))

    passed_all = 0
    total = 0

    # ── Case loop ─────────────────────────────────────────────────────────────
    for case_i, complex_prompt, steps in indexed_cases:
        row = case_i + 2
        total += 1
        print(f"\n{'='*60}", flush=True)
        print(f"CASE {total}/{len(indexed_cases)}  (level={args.level})", flush=True)
        print(f"  {complex_prompt[:120]}", flush=True)

        # Fresh OpenROAD process
        orig_dir = os.getcwd()
        run_dir = _abs(args.runDir)
        if run_dir:
            os.chdir(run_dir)
        proc = runOpenROADShell(_abs(args.openroadPath), args.loadDesignTime, slave_fd, "")
        if run_dir:
            os.chdir(orig_dir)
        time.sleep(args.loadDesignTime)
        clearQueue(oq)

        accepted_codes = []
        step_results = []
        step_outputs = []
        step_codes = []
        step_rag_contexts = []

        for step_idx, raw_prompt in enumerate(steps):
            step_n = step_idx + 1

            if not raw_prompt:
                step_results.append(True)
                step_outputs.append("")
                step_codes.append("")
                step_rag_contexts.append("")
                continue

            print(f"\n  ── Step {step_n}/{len(steps)}: {raw_prompt[:100]}", flush=True)

            # Skip if prior step failed
            if step_idx > 0 and not step_results[-1]:
                print(f"  [Step {step_n}] SKIP — step {step_n-1} failed", flush=True)
                step_results.append(False)
                step_outputs.append("")
                step_codes.append("")
                step_rag_contexts.append("")
                continue

            # Build task with prior context
            if accepted_codes:
                ctx_block = "\n\n".join(
                    f"# --- Step {i+1} (already executed) ---\n{c}"
                    for i, c in enumerate(accepted_codes)
                )
                task = (
                    f"{raw_prompt}\n\n"
                    f"[Context] The following steps have already executed in this "
                    f"OpenROAD session:\n"
                    f"```python\n{ctx_block}\n```\n\n"
                    f"Re-acquire objects you need (e.g. block = design.getBlock()) "
                    f"rather than relying on variables from previous steps."
                )
            else:
                task = raw_prompt

            # ── RAG retrieval ─────────────────────────────────────────────
            rag_ctx = retrieve_rag(task, embeddings, embed_model,
                                   documents, metadata)
            step_rag_contexts.append(rag_ctx)

            # ── Build prompt ──────────────────────────────────────────────
            user_msg = f"Task: {task}\n"
            if rag_ctx:
                user_msg += f"\n[Relevant OpenROAD API Documentation]\n{rag_ctx}\n"
            if code_examples_str:
                user_msg += f"\n{code_examples_str}\n"
            user_msg += (
                "\nWrite the complete Python script. "
                "Print all results so they are visible in the shell output."
            )

            # ── Generate (single shot) ────────────────────────────────────
            raw_output = _call_openai(
                messages=[
                    {"role": "system", "content": _SYSTEM_PROMPT},
                    {"role": "user",   "content": user_msg},
                ],
                api_key=args.openaiKey, model=args.model,
            )
            code = _extract_code(raw_output)
            step_codes.append(code)

            if not code:
                print(f"  [Step {step_n}] No code generated", flush=True)
                step_results.append(False)
                step_outputs.append("No code generated")
                continue

            # ── Execute in OpenROAD ───────────────────────────────────────
            stdout, has_tb = "", True
            try:
                cmd = processCodeString(code)
                stdout, has_tb = sendCommandOpenROAD(
                    proc, cmd, oq, args.maxWaitTime, args.commandFlushTime,
                )
            except Exception as exc:
                stdout = f"[OpenROAD error] {exc}"
                has_tb = True

            passed = not has_tb
            step_results.append(passed)
            step_outputs.append(stdout)

            if passed:
                accepted_codes.append(code)
            print(f"  [Step {step_n}] {'PASS' if passed else 'FAIL'}", flush=True)
            if not passed and stdout:
                print(f"    {stdout.strip()[:200]}", flush=True)

        # Terminate OpenROAD
        try:
            proc.terminate()
            proc.wait()
        except Exception:
            pass

        all_passed = all(step_results)
        if all_passed:
            passed_all += 1

        overall = "PASS" if all_passed else (
            f"STEP{next(i for i, r in enumerate(step_results) if not r)+1}_FAIL"
        )
        print(f"  [Overall] {overall}", flush=True)

        # ── Write Excel ───────────────────────────────────────────────────
        ws.cell(row=row, column=1, value=_clean(complex_prompt))
        for step_idx, raw_prompt in enumerate(steps):
            step_n = step_idx + 1
            ws.cell(row=row, column=_col(step_n, "prompt"),
                    value=raw_prompt)
            ws.cell(row=row, column=_col(step_n, "rag_context"),
                    value=_clean(step_rag_contexts[step_idx][:2000]) if step_idx < len(step_rag_contexts) else "")
            ws.cell(row=row, column=_col(step_n, "code"),
                    value=_clean(step_codes[step_idx]) if step_idx < len(step_codes) else "")
            ws.cell(row=row, column=_col(step_n, "result"),
                    value=("PASS" if step_results[step_idx] else "FAIL") if step_idx < len(step_results) else "")
            ws.cell(row=row, column=_col(step_n, "output"),
                    value=_clean(step_outputs[step_idx][:2000]) if step_idx < len(step_outputs) else "")
        ws.cell(row=row, column=_overall_col(), value=overall)
        wb.save(result_file)

    print(f"\n{'='*60}", flush=True)
    print(f"All steps passed: {passed_all}/{total} = "
          f"{passed_all/max(total,1)*100:.1f}%", flush=True)
    print(f"Results → {result_file}", flush=True)


if __name__ == "__main__":
    main()
